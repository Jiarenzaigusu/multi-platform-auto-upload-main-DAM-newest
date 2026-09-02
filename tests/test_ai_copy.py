"""tests.test_ai_copy 模块：AI 文案功能的单元测试。

覆盖：
- 卖点 Excel 解析与目录存储（上传/解析/删除/TTL/LRU）
- 商品链接读取（京东/天猫/通用 HTML/自定义服务，含 SSRF 防护）
- LLM 文案生成（system prompt、工具调用、字数校验、高风险表述/无依据数字校验）
- FastAPI 路由（/api/ai-copy/*）
"""
from __future__ import annotations

import asyncio
from contextlib import nullcontext
from copy import deepcopy
from io import BytesIO
import json
import unittest
from unittest.mock import Mock, patch
from urllib.parse import unquote

import certifi
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from starlette.datastructures import UploadFile

from webapp.ai_copy.contracts import (
    GeneratedCopyDraft,
    GenerateCopyRequest,
    ProductReference,
    ProductReferencesRequest,
    ProductSearchConfig,
    SellingPointReference,
)
from webapp.ai_copy.errors import (
    LLMResponseError,
    ProductLookupError,
    SellingPointCatalogError,
)
from webapp.ai_copy.product_lookup import ProductSearchTool
from webapp.ai_copy.product_lookup.cache import ProductReferenceCache
from webapp.ai_copy.product_lookup.custom_reader import _NoRedirectHandler
from webapp.ai_copy.product_lookup.generic_reader import GenericHtmlProductReader
from webapp.ai_copy.product_lookup.jd_client import (
    JD_MOBILE_HEADERS,
    PatchrightJdPageFetcher,
    _TransientJdRequestError,
)
from webapp.ai_copy.product_lookup.jd_reader import JdProductReader, extract_jd_sku
from webapp.ai_copy.product_lookup.public_http import (
    FetchedPage,
    PublicPageHttpClient,
    create_trusted_ssl_context,
    validate_public_product_url,
)
from webapp.ai_copy.product_lookup.tmall_client import DirectoryTmallStorageStateProvider
from webapp.ai_copy.product_lookup.tmall_reader import (
    TmallProductReader,
    extract_tmall_product_ids,
)
from webapp.ai_copy.router import _import_copy_to_excel, create_ai_copy_router
from webapp.ai_copy.service import AiCopyService
from webapp.ai_copy.selling_points import SellingPointCatalogStore
from webapp.ai_copy.settings import AiCopySettings


class FakeChatProvider:
    model = "test-copy-model"
    provider_label = "Test Provider"

    def __init__(
        self,
        *,
        tool_url: str = "https://shop.example/product/42",
        tool_urls: list[str] | None = None,
        ready: bool = True,
        draft: dict[str, str | list[str]] | None = None,
    ) -> None:
        self.tool_urls = tool_urls or [tool_url]
        self.ready = ready
        self.draft = draft or {
            "title": "轻盈入夏，通勤与周末自在相伴，写下舒展轻松的日常节奏呀呀",
            "body": "轻盈透气的日常鞋款，陪伴通勤、散步与周末出游，让每一次搭配都自然舒展。" * 30,
        }
        self.calls: list[dict] = []

    def chat(self, messages, **options):
        self.calls.append({"messages": deepcopy(messages), **options})
        if options.get("tools"):
            return {
                "role": "assistant",
                "content": None,
                "tool_calls": [
                    {
                        "id": f"call-product-{index}",
                        "type": "function",
                        "function": {
                            "name": "inspect_product_link",
                            "arguments": json.dumps({"url": product_url}),
                        },
                    }
                    for index, product_url in enumerate(self.tool_urls, start=1)
                ],
            }
        return {
            "role": "assistant",
            "content": json.dumps(self._candidate_payload(), ensure_ascii=False),
        }

    def _candidate_payload(self) -> dict[str, list[str]]:
        """Allow single-result fixtures while exercising the source candidate contract."""
        if "titles" in self.draft and "bodies" in self.draft:
            return self.draft  # type: ignore[return-value]
        return {
            "titles": [str(self.draft["title"])],
            "bodies": [str(self.draft["body"])],
        }

    @staticmethod
    def session():
        return nullcontext()


class FakeProductTool:
    def __init__(self) -> None:
        self.calls: list[tuple[str, ProductSearchConfig]] = []

    def inspect(self, url: str, config: ProductSearchConfig) -> ProductReference:
        self.calls.append((url, config))
        return ProductReference(
            source_url=url,
            title="轻量透气休闲鞋",
            summary="网布鞋面，适合通勤与日常步行。",
            attributes={"颜色": "米白", "尺码": "35-40"},
        )


def build_selling_point_workbook(rows: list[tuple[object, object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["商品ID或货号", "商品核心内容卖点"])
    for row in rows:
        worksheet.append(list(row))
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class SellingPointCatalogTests(unittest.TestCase):
    def test_upload_and_resolve_multiple_unique_identifiers(self):
        store = SellingPointCatalogStore()
        uploaded = store.upload(
            "商品核心卖点.xlsx",
            build_selling_point_workbook(
                [
                    ("SKU-001", "轻量透气，适合夏日通勤"),
                    ("000002", "柔软好搭，适合日常穿着"),
                ]
            ),
        )

        resolved = store.resolve(uploaded.catalog_id, ["sku-001", "000002"])

        self.assertEqual(uploaded.row_count, 2)
        self.assertEqual([item.identifier for item in resolved], ["SKU-001", "000002"])
        self.assertIn("夏日通勤", resolved[0].selling_point)

    def test_duplicate_identifier_is_rejected_case_insensitively(self):
        store = SellingPointCatalogStore()

        with self.assertRaisesRegex(SellingPointCatalogError, "重复"):
            store.upload(
                "points.xlsx",
                build_selling_point_workbook(
                    [("SKU-001", "卖点一"), ("sku-001", "卖点二")]
                ),
            )

    def test_unknown_identifier_is_rejected(self):
        store = SellingPointCatalogStore()
        uploaded = store.upload(
            "points.xlsx",
            build_selling_point_workbook([("SKU-001", "卖点一")]),
        )

        with self.assertRaisesRegex(SellingPointCatalogError, "SKU-404"):
            store.resolve(uploaded.catalog_id, ["SKU-404"])


class AiCopyServiceTests(unittest.TestCase):
    @staticmethod
    def make_request(
        service: AiCopyService,
        selling_point: str,
        **values,
    ) -> GenerateCopyRequest:
        identifier = values.pop("identifier", "SKU-001")
        catalog = service.upload_selling_points(
            "points.xlsx",
            build_selling_point_workbook([(identifier, selling_point)]),
        )
        return GenerateCopyRequest(
            selling_point_catalog_id=catalog.catalog_id,
            product_identifiers=[identifier],
            style=values.pop("style", "atmospheric_seeding"),
            scene=values.pop("scene", "daily_styling"),
            **values,
        )

    def test_product_link_is_read_through_required_llm_tool_call(self):
        provider = FakeChatProvider()
        product_tool = FakeProductTool()
        service = AiCopyService(provider, product_tool)
        request = self.make_request(
            service,
            "突出轻便、透气和通勤百搭",
            festival="情人节",
            product_urls=["https://shop.example/product/42"],
            product_search={
                "endpoint_url": "https://search.example/inspect",
                "api_key": "request-only-secret",
            },
        )

        result = service.generate(request)

        self.assertEqual(result.model, "test-copy-model")
        self.assertEqual(result.provider, "Test Provider")
        self.assertEqual(result.product_references[0].title, "轻量透气休闲鞋")
        self.assertEqual(len(provider.calls), 2)
        self.assertEqual(provider.calls[0]["tool_choice"]["function"]["name"], "inspect_product_link")
        self.assertEqual(product_tool.calls[0][0], "https://shop.example/product/42")
        self.assertEqual(product_tool.calls[0][1].api_key, "request-only-secret")
        final_messages = provider.calls[1]["messages"]
        self.assertEqual(final_messages[-2]["role"], "tool")
        self.assertIn("轻量透气休闲鞋", final_messages[-2]["content"])

    def test_generation_without_product_link_does_not_run_product_tool(self):
        provider = FakeChatProvider()
        product_tool = FakeProductTool()
        service = AiCopyService(provider, product_tool)

        result = service.generate(
            self.make_request(
                service,
                "夏季轻量通勤鞋",
                style="relaxed_natural",
                scene="daily_styling",
            )
        )

        self.assertEqual(result.product_references, [])
        self.assertEqual(product_tool.calls, [])
        self.assertEqual(len(provider.calls), 1)
        self.assertNotIn("tools", provider.calls[0])

    def test_copy_prompt_requires_grounded_and_non_absolute_language(self):
        messages = AiCopyService._initial_messages(
            self.make_request(
                AiCopyService(FakeChatProvider(), FakeProductTool()),
                "不要使用功效词汇",
                style="relaxed_natural",
                scene="daily_styling",
            ),
            [SellingPointReference(identifier="SKU-001", selling_point="不要使用功效词汇")],
        )

        system_prompt = messages[0]["content"]
        self.assertIn("阿拉伯数字", system_prompt)
        self.assertIn("功效", system_prompt)
        self.assertIn("100%", system_prompt)

    def test_custom_copy_context_overrides_the_preset_in_the_prompt(self):
        request = GenerateCopyRequest(
            selling_point_catalog_id="a" * 16,
            product_identifiers=["SKU-001"],
            style=None,
            scene=None,
            festival=None,
            custom_style="松弛的法式随笔",
            custom_scene="雨天咖啡馆约会",
            custom_festival="品牌周年庆",
        )

        messages = AiCopyService._initial_messages(
            request,
            [SellingPointReference(identifier="SKU-001", selling_point="轻量透气，适合日常通勤")],
        )

        prompt = messages[1]["content"]
        self.assertIn("文案风格：松弛的法式随笔", prompt)
        self.assertIn("内容场景：雨天咖啡馆约会", prompt)
        self.assertIn("节日氛围：品牌周年庆", prompt)

    def test_llm_cannot_change_the_requested_product_url(self):
        provider = FakeChatProvider(tool_url="https://attacker.example/private")
        service = AiCopyService(provider, FakeProductTool())

        with self.assertRaisesRegex(LLMResponseError, "链接与用户请求不一致"):
            service.generate(
                self.make_request(
                    service,
                    "商品介绍",
                    style="old_money_luxury",
                    scene="daily_styling",
                    product_urls=["https://shop.example/product/42"],
                )
            )

    def test_high_risk_claim_is_rejected_after_generation(self):
        provider = FakeChatProvider(
            draft={"title": "销量第一的选择", "body": "轻松搭配日常造型，通勤穿着也很自在。"}
        )
        service = AiCopyService(provider, FakeProductTool())

        with self.assertRaisesRegex(LLMResponseError, "高风险"):
            service.generate(
                self.make_request(service, "日常百搭休闲鞋")
            )

    def test_invented_numeric_claim_is_rejected(self):
        provider = FakeChatProvider(
            draft={"title": "轻盈日常鞋", "body": "采用99%轻量设计，通勤更自在。"}
        )
        service = AiCopyService(provider, FakeProductTool())

        with self.assertRaisesRegex(LLMResponseError, "没有的数字信息：99%"):
            service.generate(
                self.make_request(service, "突出轻便和通勤百搭")
            )

    def test_source_grounded_number_is_allowed(self):
        provider = FakeChatProvider(
            draft={"title": "轻盈日常鞋", "body": "鞋面含棉99%，通勤穿着自然舒适。"}
        )
        service = AiCopyService(provider, FakeProductTool())

        result = service.generate(
            self.make_request(
                service,
                "鞋面含棉99%，适合日常通勤",
                title_max_chars=5,
                body_max_chars=19,
            )
        )

        self.assertIn("99%", result.body)

    def test_product_url_number_does_not_ground_a_marketing_claim(self):
        provider = FakeChatProvider(
            tool_url="https://shop.example/product/99",
            draft={"title": "轻盈日常鞋", "body": "99%用户都会喜欢的通勤选择。"},
        )
        service = AiCopyService(provider, FakeProductTool())

        with self.assertRaisesRegex(LLMResponseError, "没有的数字信息：99%"):
            service.generate(
                self.make_request(
                    service,
                    "突出轻便和通勤百搭",
                    product_urls=["https://shop.example/product/99"],
                )
            )

    def test_multiple_product_links_are_all_read_before_generation(self):
        product_urls = [
            "https://shop.example/product/alpha",
            "https://shop.example/product/beta",
        ]
        provider = FakeChatProvider(tool_urls=product_urls)
        product_tool = FakeProductTool()
        service = AiCopyService(provider, product_tool)

        result = service.generate(
            self.make_request(
                service,
                "适合日常搭配与通勤",
                product_urls=product_urls,
            )
        )

        self.assertEqual(
            [reference.source_url for reference in result.product_references],
            product_urls,
        )
        self.assertEqual([call[0] for call in product_tool.calls], product_urls)
        self.assertEqual(len(provider.calls[0]["messages"]), 2)
        self.assertEqual(len(provider.calls[0]["tools"]), 1)
        for product_url in product_urls:
            self.assertIn(product_url, provider.calls[0]["messages"][1]["content"])
        final_messages = provider.calls[1]["messages"]
        self.assertEqual(
            [message["role"] for message in final_messages[-3:-1]],
            ["tool", "tool"],
        )
        self.assertTrue(
            all(
                "轻量透气休闲鞋" in message["content"]
                for message in final_messages[-3:-1]
            )
        )

    def test_product_links_are_deduplicated_and_limited(self):
        service = AiCopyService(FakeChatProvider(), FakeProductTool())
        product_url = "https://shop.example/product/alpha"

        request = self.make_request(
            service,
            "日常百搭",
            product_urls=[product_url, product_url],
        )

        self.assertEqual(
            [str(value) for value in request.product_urls],
            [product_url],
        )
        with self.assertRaisesRegex(ValueError, "20 个商品链接"):
            self.make_request(
                service,
                "日常百搭",
                product_urls=[
                    f"https://shop.example/product/item-{index}"
                    for index in range(21)
                ],
            )

    def test_multiple_selling_points_are_sent_as_important_references(self):
        provider = FakeChatProvider()
        service = AiCopyService(provider, FakeProductTool())
        catalog = service.upload_selling_points(
            "points.xlsx",
            build_selling_point_workbook(
                [
                    ("SKU-001", "轻量透气，适合通勤"),
                    ("SKU-002", "柔软百搭，适合周末出游"),
                ]
            ),
        )

        result = service.generate(
            GenerateCopyRequest(
                selling_point_catalog_id=catalog.catalog_id,
                product_identifiers=["SKU-001", "SKU-002"],
                style="atmospheric_seeding",
                scene="daily_styling",
            )
        )

        prompt = provider.calls[0]["messages"][1]["content"]
        self.assertIn("SKU-001：轻量透气", prompt)
        self.assertIn("SKU-002：柔软百搭", prompt)
        self.assertIn("标题和正文的参考，生成的标题文案结果中引用该核心卖点的文字占比约50%", prompt)
        self.assertEqual(len(result.selling_point_references), 2)

    def test_direct_copy_reference_keeps_equal_weight_with_selling_points(self):
        provider = FakeChatProvider()
        service = AiCopyService(provider, FakeProductTool())

        service.generate(
            self.make_request(
                service,
                "轻量透气，适合日常通勤",
                copy_reference=" 松弛自然的日常穿搭叙述 ",
            )
        )

        prompt = provider.calls[0]["messages"][1]["content"]
        self.assertIn("文案参考（用户直接输入", prompt)
        self.assertIn("与核心卖点各占生成内容约50%的权重", prompt)
        self.assertIn("松弛自然的日常穿搭叙述", prompt)

    def test_manual_selling_point_is_used_without_excel_catalog(self):
        provider = FakeChatProvider()
        service = AiCopyService(provider, FakeProductTool())

        result = service.generate(
            GenerateCopyRequest(
                selling_point_input_mode="manual",
                manual_selling_point="轻盈透气，适合日常通勤",
                style="atmospheric_seeding",
                scene="daily_styling",
            )
        )

        prompt = provider.calls[0]["messages"][1]["content"]
        self.assertIn("来自用户直接输入", prompt)
        self.assertIn("轻盈透气，适合日常通勤", prompt)
        self.assertNotIn("来自用户上传 Excel", prompt)
        self.assertEqual(result.selling_point_references[0].identifier, "直接输入")

    def test_selling_point_input_modes_reject_mixed_fields(self):
        with self.assertRaisesRegex(ValueError, "直接输入模式不能同时提交"):
            GenerateCopyRequest(
                selling_point_input_mode="manual",
                selling_point_catalog_id="a" * 16,
                product_identifiers=["SKU-001"],
                manual_selling_point="轻盈透气",
                style="atmospheric_seeding",
                scene="daily_styling",
            )

        with self.assertRaisesRegex(ValueError, "Excel 模式不能同时提交"):
            GenerateCopyRequest(
                selling_point_input_mode="excel",
                selling_point_catalog_id="a" * 16,
                product_identifiers=["SKU-001"],
                manual_selling_point="不应生效",
                style="atmospheric_seeding",
                scene="daily_styling",
            )

    def test_generated_copy_prompts_for_exact_han_character_targets(self):
        provider = FakeChatProvider(
            draft={"title": "夏鞋", "body": "舒适好搭，日常通勤穿着轻松。"}
        )
        service = AiCopyService(provider, FakeProductTool())

        result = service.generate(
            self.make_request(
                service,
                "轻盈透气，适合日常通勤",
                title_max_chars=15,
                body_max_chars=100,
            )
        )

        self.assertEqual(result.title, "夏鞋")
        self.assertEqual(result.body, "舒适好搭，日常通勤穿着轻松。")
        instruction = provider.calls[0]["messages"][-1]["content"]
        self.assertIn("每条标题必须正好包含 15 个汉字", instruction)
        self.assertIn("每条正文必须严格以 100 个汉字为目标", instruction)
        self.assertIn("标点、数字、英文字母、空格和其他符号全部不计", instruction)
        self.assertIn("标题达到八个汉字时", instruction)

    def test_generated_copy_returns_the_requested_candidate_counts(self):
        provider = FakeChatProvider(
            draft={
                "titles": ["轻盈通勤鞋", "自在日常鞋", "夏日好搭鞋"],
                "bodies": [
                    "轻盈透气的鞋款，适合日常通勤与自在搭配。",
                    "简约鞋型自然好搭，让出行与休闲都更轻松。",
                ],
            }
        )
        service = AiCopyService(provider, FakeProductTool())

        result = service.generate(
            self.make_request(service, "轻盈透气，适合日常通勤", title_count=3, body_count=2)
        )

        self.assertEqual(result.titles, ["轻盈通勤鞋", "自在日常鞋", "夏日好搭鞋"])
        self.assertEqual(len(result.bodies), 2)
        self.assertEqual(result.title_count, 3)
        self.assertEqual(result.body_count, 2)
        instruction = provider.calls[0]["messages"][-1]["content"]
        self.assertIn("titles 必须且只能包含 3 条不同标题", instruction)
        self.assertIn("bodies 必须且只能包含 2 条不同正文", instruction)

    def test_generated_copy_rejects_a_wrong_candidate_count(self):
        draft = GeneratedCopyDraft(
            titles=["轻盈通勤鞋", "自在日常鞋"],
            bodies=["轻盈透气的鞋款，适合日常通勤与自在搭配。"],
        )

        with self.assertRaisesRegex(LLMResponseError, "标题数量不正确"):
            AiCopyService._validate_draft_counts(draft, title_count=3, body_count=1)


class _FakeHeaders:
    @staticmethod
    def get_content_type() -> str:
        return "text/html"

    @staticmethod
    def get_content_charset() -> str:
        return "utf-8"


class _FakeResponse:
    def __init__(self, content: bytes) -> None:
        self.content = content
        self.headers = _FakeHeaders()

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return False

    def read(self, limit: int = -1) -> bytes:
        return self.content if limit < 0 else self.content[:limit]


class _FakeOpener:
    def __init__(self, response: _FakeResponse) -> None:
        self.response = response

    def open(self, *_args, **_kwargs) -> _FakeResponse:
        return self.response


class ProductSearchToolTests(unittest.TestCase):
    def setUp(self) -> None:
        self.settings = AiCopySettings()

    def test_custom_search_service_uses_strict_contract_and_request_key(self):
        response = _FakeResponse(
            json.dumps(
                {
                    "title": "夏季凉感床品",
                    "summary": "柔软亲肤，适合夏季卧室。",
                    "attributes": {"规格": "四件套"},
                },
                ensure_ascii=False,
            ).encode()
        )
        opener = Mock()
        opener.open.return_value = response
        with patch(
            "webapp.ai_copy.product_lookup.custom_reader.build_opener",
            return_value=opener,
        ):
            result = ProductSearchTool(self.settings).inspect(
                "https://shop.example/item/1",
                ProductSearchConfig(
                    endpoint_url="https://search.example/inspect",
                    api_key="one-request-key",
                ),
            )

        sent_request = opener.open.call_args.args[0]
        self.assertEqual(sent_request.get_header("Authorization"), "Bearer one-request-key")
        self.assertEqual(json.loads(sent_request.data), {"url": "https://shop.example/item/1"})
        self.assertEqual(result.attributes, {"规格": "四件套"})

    def test_custom_search_key_requires_https_endpoint(self):
        with self.assertRaisesRegex(ValueError, "HTTPS"):
            ProductSearchConfig(
                endpoint_url="http://search.example/inspect",
                api_key="one-request-key",
            )

    def test_custom_service_disables_redirects_that_could_forward_its_key(self):
        response = _FakeResponse(
            json.dumps(
                {"title": "商品", "summary": "商品摘要", "attributes": {}},
                ensure_ascii=False,
            ).encode()
        )
        opener = Mock()
        opener.open.return_value = response
        with patch(
            "webapp.ai_copy.product_lookup.custom_reader.build_opener",
            return_value=opener,
        ) as build:
            ProductSearchTool(self.settings).inspect(
                "https://shop.example/item/1",
                ProductSearchConfig(
                    endpoint_url="https://search.example/inspect",
                    api_key="one-request-key",
                ),
            )

        self.assertIsInstance(build.call_args.args[0], _NoRedirectHandler)

    def test_public_page_retries_each_validated_address(self):
        attempts: list[str] = []

        class FakeConnection:
            def __init__(self, _host, _port, address, _timeout, _ssl_context):
                self.address = address

            def request(self, *_args, **_kwargs):
                attempts.append(self.address)
                if self.address == "203.0.113.10":
                    raise OSError("unreachable")

            @staticmethod
            def getresponse():
                response = _FakeResponse(b"<html></html>")
                response.status = 200
                return response

            @staticmethod
            def close():
                return None

        with patch(
            "webapp.ai_copy.product_lookup.public_http.validate_public_product_url",
            return_value=["203.0.113.10", "203.0.113.11"],
        ), patch(
            "webapp.ai_copy.product_lookup.public_http._PinnedHTTPSConnection",
            FakeConnection,
        ):
            page = PublicPageHttpClient(
                timeout_seconds=1, max_bytes=1000, ssl_context=Mock()
            ).get("https://shop.example/item/1")

        self.assertEqual(attempts, ["203.0.113.10", "203.0.113.11"])
        self.assertEqual(
            (page.content, page.content_type, page.charset),
            (b"<html></html>", "text/html", "utf-8"),
        )

    def test_public_url_rejects_embedded_credentials(self):
        with self.assertRaisesRegex(ProductLookupError, "公开"):
            validate_public_product_url("https://user:password@shop.example/item/1")

    def test_ssl_context_uses_certifi_ca_bundle(self):
        context = Mock()
        with patch(
            "webapp.ai_copy.product_lookup.public_http.ssl.create_default_context",
            return_value=context,
        ) as create_context:
            result = create_trusted_ssl_context()

        self.assertIs(result, context)
        create_context.assert_called_once_with(cafile=certifi.where())

    def test_public_page_extracts_product_json_ld(self):
        page = b"""
        <html><head><script type="application/ld+json">
        {"@type":"Product","name":"Everyday Sneaker","description":"Light mesh upper",
         "brand":{"name":"North Star"},"sku":"NS-42",
         "offers":{"price":"399","priceCurrency":"CNY"}}
        </script></head></html>
        """
        http_client = Mock()
        http_client.get.return_value = FetchedPage(
            page, "text/html", "utf-8", "https://shop.example/item/42"
        )
        result = GenericHtmlProductReader(http_client).inspect(
            "https://shop.example/item/42"
        )

        self.assertEqual(result.title, "Everyday Sneaker")
        self.assertEqual(result.attributes["品牌"], "North Star")
        self.assertEqual(result.attributes["价格"], "399 CNY")

    def test_public_page_rejects_private_dns_results(self):
        with patch(
            "webapp.ai_copy.product_lookup.public_http.socket.getaddrinfo",
            return_value=[(2, 1, 6, "", ("127.0.0.1", 443))],
        ):
            with self.assertRaisesRegex(ProductLookupError, "内网"):
                validate_public_product_url("https://shop.example/item/42")

    def test_jd_reader_uses_mobile_page_and_extracts_structured_product(self):
        page = """
        <script>
        window._itemOnly = ({
          "item": {
            "skuName": "盖璞儿童连帽抓绒卫衣 ...",
            "brandName": "盖璞（GAP）",
            "spAttr": {
              "product_features": "nameWithoutBrand:儿童连帽抓绒卫衣;descModel:1"
            },
            "saleProp": {"1": "颜色", "2": "尺码"},
            "salePropSeq": {
              "1": ["黄色", "灰色"],
              "2": ["120 cm", "130 cm"]
            },
            "newColorSize": [
              {"skuId": "10230424567386", "color": "灰色", "size": "120 cm"}
            ]
          }
        });
        window._itemInfo = ({
          "product": {}
        });
        </script>
        """.encode()
        http_client = Mock()
        http_client.get.return_value = FetchedPage(
            page,
            "text/html",
            "utf-8",
            "https://item.m.jd.com/product/10230424567386.html",
        )

        reader = JdProductReader(http_client)
        result = reader.inspect(
            "https://item.jd.com/10230424567386.html?sdx=tracking"
        )

        http_client.get.assert_called_once_with(
            "https://item.m.jd.com/product/10230424567386.html",
            headers=JD_MOBILE_HEADERS,
        )
        self.assertEqual(result.source_url, "https://item.jd.com/10230424567386.html?sdx=tracking")
        self.assertEqual(
            result.title, "盖璞（GAP） 儿童连帽抓绒卫衣 灰色 120 cm"
        )
        self.assertEqual(result.attributes, {})

        cached = reader.inspect(
            "https://item.jd.com/10230424567386.html?sdx=another-request"
        )
        self.assertEqual(
            cached.source_url,
            "https://item.jd.com/10230424567386.html?sdx=another-request",
        )
        self.assertEqual(http_client.get.call_count, 1)

    def test_jd_reader_uses_recent_success_when_refresh_is_limited(self):
        now = [0.0]
        cache = ProductReferenceCache(
            fresh_seconds=10,
            stale_seconds=100,
            clock=lambda: now[0],
        )
        page_fetcher = Mock()
        page_fetcher.get.return_value = FetchedPage(
            b'<script>window._itemOnly=({"item":{"skuName":"JD Product"}})</script>',
            "text/html",
            "utf-8",
            "https://item.m.jd.com/product/42.html",
        )
        reader = JdProductReader(page_fetcher, cache)
        first = reader.inspect("https://item.jd.com/42.html")
        now[0] = 11
        page_fetcher.get.side_effect = ProductLookupError("京东商品页面限制了当前读取请求")

        fallback = reader.inspect("https://item.jd.com/42.html?retry=1")

        self.assertEqual(first.title, "JD Product")
        self.assertEqual(fallback.title, "JD Product")
        self.assertEqual(fallback.source_url, "https://item.jd.com/42.html?retry=1")

    def test_jd_client_retries_transient_limit_response(self):
        expected = FetchedPage(
            b"<html></html>",
            "text/html",
            "utf-8",
            "https://item.m.jd.com/product/42.html",
        )
        fetcher = PatchrightJdPageFetcher(
            timeout_seconds=1,
            max_bytes=1000,
            max_attempts=3,
            retry_base_seconds=0.1,
        )
        with patch.object(
            fetcher,
            "_get_once",
            side_effect=[_TransientJdRequestError("limited"), expected],
        ) as get_once, patch(
            "webapp.ai_copy.product_lookup.jd_client.time.sleep"
        ) as sleep:
            result = fetcher.get("https://item.m.jd.com/product/42.html")

        self.assertIs(result, expected)
        self.assertEqual(get_once.call_count, 2)
        sleep.assert_called_once_with(0.1)

    def test_jd_reader_recognizes_desktop_and_mobile_links(self):
        self.assertEqual(
            extract_jd_sku("https://item.jd.com/10230424567386.html?foo=bar"),
            "10230424567386",
        )
        self.assertEqual(
            extract_jd_sku(
                "https://item.m.jd.com/product/10230424567386.html"
            ),
            "10230424567386",
        )
        self.assertIsNone(extract_jd_sku("https://example.com/10230424567386.html"))

    def test_tmall_reader_extracts_selected_sku_and_product_facts(self):
        response = {
            "item": {
                "itemId": "1006533002222",
                "title": "Gap 纯棉宽松短袖 T 恤",
            },
            "skuBase": {
                "skus": [
                    {
                        "skuId": "6003757841492",
                        "propPath": "1627207:43948691702;20509:382156294",
                    }
                ],
                "props": [
                    {
                        "pid": "1627207",
                        "name": "颜色",
                        "valueMap": {
                            "43948691702": {"name": "淡粉色729157"}
                        },
                    },
                    {
                        "pid": "20509",
                        "name": "尺码",
                        "valueMap": {
                            "382156294": {"name": "170/92A(M) 亚洲尺码"}
                        },
                    },
                ],
            },
            "plusViewVO": {
                "industryParamVO": {
                    "enhanceParamList": [
                        {"propertyName": "材质成分", "valueName": "棉100%"},
                        {"propertyName": "版型分类", "valueName": "宽松型"},
                    ],
                    "basicParamList": [
                        {"propertyName": "品牌", "valueName": "Gap"},
                        {"propertyName": "颜色", "valueName": "不应读取全部颜色"},
                    ],
                }
            },
        }
        payload = {
            "appData": None,
            "loaderData": {"home": {"data": {"res": response}}},
        }
        page_fetcher = Mock()
        page_fetcher.get.return_value = FetchedPage(
            f"<script>var b = {json.dumps(payload, ensure_ascii=False)}</script>".encode(),
            "text/html",
            "utf-8",
            "https://detail.tmall.com/item.htm?id=1006533002222",
        )
        reader = TmallProductReader(page_fetcher)
        product_url = (
            "https://detail.tmall.com/item.htm?id=1006533002222"
            "&skuId=6003757841492&spm=tracking"
        )

        result = reader.inspect(product_url)

        self.assertEqual(result.source_url, product_url)
        self.assertEqual(result.title, "Gap 纯棉宽松短袖 T 恤")
        self.assertIn("当前颜色：淡粉色729157", result.summary)
        self.assertIn("当前尺码：170/92A(M) 亚洲尺码", result.summary)
        self.assertIn("品牌：Gap", result.summary)
        self.assertIn("材质成分：棉100%", result.summary)
        self.assertNotIn("不应读取全部颜色", result.summary)
        self.assertEqual(result.attributes, {})

        cached = reader.inspect(
            "https://detail.tmall.com/item.htm?id=1006533002222"
            "&skuId=6003757841492&spm=another"
        )
        self.assertEqual(page_fetcher.get.call_count, 1)
        self.assertIn("spm=another", cached.source_url)

    def test_tmall_reader_rejects_login_page(self):
        page_fetcher = Mock()
        page_fetcher.get.return_value = FetchedPage(
            b'<script>window._config_={"action":"login"}</script>',
            "text/html",
            "utf-8",
            "https://login.taobao.com/member/login.jhtml",
        )

        with self.assertRaisesRegex(ProductLookupError, "没有返回可解析"):
            TmallProductReader(page_fetcher).inspect(
                "https://detail.tmall.com/item.htm?id=1006533002222"
            )

    def test_tmall_reader_recognizes_desktop_and_mobile_links(self):
        self.assertEqual(
            extract_tmall_product_ids(
                "https://detail.tmall.com/item.htm?id=1006533002222"
                "&skuId=6003757841492"
            ),
            ("1006533002222", "6003757841492"),
        )
        self.assertEqual(
            extract_tmall_product_ids(
                "https://detail.m.tmall.com/item.htm?id=1006533002222"
            ),
            ("1006533002222", None),
        )
        self.assertIsNone(
            extract_tmall_product_ids(
                "https://detail.tmall.com/item.htm?id=invalid&skuId=123"
            )
        )
        self.assertIsNone(
            extract_tmall_product_ids(
                "https://example.com/item.htm?id=1006533002222"
            )
        )

    def test_tmall_storage_provider_finds_account_cookie_files(self):
        import tempfile
        from pathlib import Path

        state = {
            "cookies": [
                {
                    "name": "_m_h5_tk",
                    "value": "token",
                    "domain": ".tmall.com",
                }
            ]
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            cookie_dir = Path(temp_dir)
            current = cookie_dir / "shop1.json"
            legacy = cookie_dir / "tmall_shop2.json"
            ignored = cookie_dir / "jd_shop.json"
            current.write_text(json.dumps(state), encoding="utf-8")
            legacy.write_text(json.dumps(state), encoding="utf-8")
            ignored.write_text(json.dumps({"cookies": []}), encoding="utf-8")

            candidates = DirectoryTmallStorageStateProvider(
                cookie_dir, max_candidates=5
            ).candidates()

        self.assertIn(current.resolve(), candidates)
        self.assertIn(legacy.resolve(), candidates)
        self.assertNotIn(ignored.resolve(), candidates)


class AiCopyRouterTests(unittest.TestCase):
    def test_selling_point_template_download_matches_upload_contract(self):
        service = AiCopyService(FakeChatProvider(), FakeProductTool())
        router = create_ai_copy_router(service)
        endpoint = next(
            route.endpoint
            for route in router.routes
            if route.path == "/api/ai-copy/selling-point-template"
        )

        response = endpoint()

        self.assertIsInstance(response, StreamingResponse)
        self.assertEqual(
            response.media_type,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        disposition = response.headers["content-disposition"]
        self.assertIn('filename="selling-point-template.xlsx"', disposition)
        encoded_filename = disposition.split("filename*=UTF-8''", 1)[1]
        self.assertEqual(unquote(encoded_filename), "商品核心卖点模板.xlsx")

        async def read_response_body() -> bytes:
            chunks = []
            async for chunk in response.body_iterator:
                chunks.append(bytes(chunk))
            return b"".join(chunks)

        workbook = load_workbook(BytesIO(asyncio.run(read_response_body())))
        try:
            self.assertEqual(workbook.sheetnames, ["商品核心卖点"])
            worksheet = workbook["商品核心卖点"]
            self.assertEqual(
                [worksheet["A1"].value, worksheet["B1"].value],
                ["商品ID或货号", "商品核心内容卖点"],
            )
            self.assertEqual(worksheet["A2"].value, "SKU-001")
            self.assertIn("日常通勤", worksheet["B2"].value)
            self.assertEqual(worksheet.freeze_panes, "A2")
        finally:
            workbook.close()

    def test_import_copy_matches_a_multi_id_group_regardless_of_separator_or_order(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["视频路径", "标题", "文案", "商品ID"])
        worksheet.append([
            "photos/demo.mp4",
            "旧标题",
            "旧文案",
            "1049855907469，1061612282569",
        ])
        content = BytesIO()
        workbook.save(content)
        workbook.close()

        modified_content, stats = _import_copy_to_excel(
            content.getvalue(),
            "新标题",
            "新文案",
            ["1061612282569", "1049855907469"],
        )

        imported = load_workbook(BytesIO(modified_content))
        try:
            worksheet = imported.active
            self.assertEqual(worksheet["B2"].value, "新标题")
            self.assertEqual(worksheet["C2"].value, "新文案")
            self.assertEqual(worksheet.max_row, 2)
            self.assertEqual(stats, {"matched": 1, "created": 0})
        finally:
            imported.close()

    def test_import_copy_matches_ids_separated_by_newlines_in_a_goods_cell(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["视频路径", "标题", "文案", "商品ID"])
        worksheet.append([
            "photos/demo.mp4",
            "旧标题",
            "旧文案",
            "1049855907469\n1061612282569\r\n1059642023424",
        ])
        content = BytesIO()
        workbook.save(content)
        workbook.close()

        modified_content, stats = _import_copy_to_excel(
            content.getvalue(),
            "新标题",
            "新文案",
            ["1059642023424", "1049855907469", "1061612282569"],
        )

        imported = load_workbook(BytesIO(modified_content))
        try:
            worksheet = imported.active
            self.assertEqual(worksheet["B2"].value, "新标题")
            self.assertEqual(worksheet["C2"].value, "新文案")
            self.assertEqual(stats, {"matched": 1, "created": 0})
        finally:
            imported.close()

    def test_import_copy_does_not_match_a_single_id_to_a_multi_id_row(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["视频路径", "标题", "文案", "商品ID"])
        worksheet.append([
            "photos/group.mp4",
            "多商品旧标题",
            "多商品旧文案",
            "1049855907469\n1061612282569",
        ])
        worksheet.append([
            "photos/other.mp4",
            "其他旧标题",
            "其他旧文案",
            "1059642023424",
        ])
        content = BytesIO()
        workbook.save(content)
        workbook.close()

        modified_content, stats = _import_copy_to_excel(
            content.getvalue(),
            "新标题",
            "新文案",
            ["1061612282569"],
        )

        imported = load_workbook(BytesIO(modified_content))
        try:
            worksheet = imported.active
            self.assertEqual(worksheet["B2"].value, "多商品旧标题")
            self.assertEqual(worksheet["C2"].value, "多商品旧文案")
            self.assertEqual(worksheet["B4"].value, "新标题")
            self.assertEqual(worksheet["C4"].value, "新文案")
            self.assertEqual(worksheet["D4"].value, "1061612282569")
            self.assertEqual(stats, {"matched": 0, "created": 1})
        finally:
            imported.close()

    def test_import_copy_does_not_match_an_overlapping_but_different_multi_id_group(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["视频路径", "标题", "文案", "商品ID"])
        worksheet.append([
            "photos/group.mp4",
            "原组合标题",
            "原组合文案",
            "1049855907469\n1061612282569\n1059642023424",
        ])
        content = BytesIO()
        workbook.save(content)
        workbook.close()

        modified_content, stats = _import_copy_to_excel(
            content.getvalue(),
            "新标题",
            "新文案",
            ["1049855907469", "1061612282569"],
        )

        imported = load_workbook(BytesIO(modified_content))
        try:
            worksheet = imported.active
            self.assertEqual(worksheet["B2"].value, "原组合标题")
            self.assertEqual(worksheet["B3"].value, "新标题")
            self.assertEqual(worksheet["D3"].value, "1049855907469\n1061612282569")
            self.assertEqual(stats, {"matched": 0, "created": 1})
        finally:
            imported.close()

    def test_import_copy_creates_a_row_for_an_unknown_identifier_group(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["视频路径", "标题", "文案", "商品ID"])
        worksheet.append(["photos/demo.mp4", "旧标题", "旧文案", "1049855907469"])
        content = BytesIO()
        workbook.save(content)
        workbook.close()

        modified_content, stats = _import_copy_to_excel(
            content.getvalue(),
            "新标题",
            "新文案",
            ["1061612282569", "1059642023424"],
        )

        imported = load_workbook(BytesIO(modified_content))
        try:
            worksheet = imported.active
            self.assertEqual(worksheet.max_row, 3)
            self.assertEqual(worksheet["A3"].value, None)
            self.assertEqual(worksheet["B3"].value, "新标题")
            self.assertEqual(worksheet["C3"].value, "新文案")
            self.assertEqual(worksheet["D3"].value, "1061612282569\n1059642023424")
            self.assertEqual(stats, {"matched": 0, "created": 1})
        finally:
            imported.close()

    def test_import_copy_appends_after_last_value_not_template_validation_range(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["视频路径", "标题", "文案", "商品ID", "创作者声明"])
        worksheet.append(["photos/demo.mp4", "旧标题", "旧文案", "1049855907469", "内容无需标注"])
        validation = DataValidation(type="list", formula1='"内容无需标注,内容含营销广告"')
        worksheet.add_data_validation(validation)
        validation.add("E2:E201")
        content = BytesIO()
        workbook.save(content)
        workbook.close()

        modified_content, stats = _import_copy_to_excel(
            content.getvalue(),
            "新标题",
            "新文案",
            ["1061612282569"],
        )

        imported = load_workbook(BytesIO(modified_content))
        try:
            worksheet = imported.active
            self.assertEqual(worksheet["B3"].value, "新标题")
            self.assertEqual(worksheet["C3"].value, "新文案")
            self.assertEqual(worksheet["D3"].value, "1061612282569")
            self.assertIsNone(worksheet["B202"].value)
            self.assertEqual(stats, {"matched": 0, "created": 1})
        finally:
            imported.close()

    def test_import_copy_only_writes_title_when_the_workbook_has_no_body_column(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["视频路径", "标题", "商品ID"])
        worksheet.append(["photos/demo.mp4", "旧标题", "1061612282569"])
        content = BytesIO()
        workbook.save(content)
        workbook.close()

        modified_content, stats = _import_copy_to_excel(
            content.getvalue(),
            "新标题",
            "不应写入的文案",
            ["1061612282569"],
        )

        imported = load_workbook(BytesIO(modified_content))
        try:
            worksheet = imported.active
            self.assertEqual(worksheet["B2"].value, "新标题")
            self.assertEqual(worksheet.max_column, 3)
            self.assertEqual(stats, {"matched": 1, "created": 0})
        finally:
            imported.close()

    def test_options_expose_ui_choices_and_llm_readiness(self):
        service = AiCopyService(FakeChatProvider(ready=False), FakeProductTool())
        router = create_ai_copy_router(service)
        endpoint = next(route.endpoint for route in router.routes if route.path.endswith("/options"))

        body = endpoint()

        self.assertFalse(body["llm"]["ready"])
        self.assertEqual(body["llm"]["model"], "test-copy-model")
        self.assertEqual(body["llm"]["provider"], "Test Provider")
        self.assertEqual(
            body["styles"],
            [
                {"value": "old_money_luxury", "label": "老钱轻奢"},
                {"value": "relaxed_natural", "label": "自然松弛"},
                {"value": "gentle_healing", "label": "治愈温柔"},
                {"value": "retro_atmosphere", "label": "复古氛围"},
                {"value": "atmospheric_seeding", "label": "氛围感种草"},
                {"value": "sweet_cute", "label": "甜美可爱"},
                {"value": "cool_bold", "label": "酷感飒爽"},
            ],
        )
        self.assertEqual(
            body["scenes"],
            [
                {"value": "fitness", "label": "运动健身"},
                {"value": "parenting_and_baby", "label": "母婴亲子"},
                {"value": "leisure_travel", "label": "度假休闲出游"},
                {"value": "daily_styling", "label": "日常穿搭"},
                {"value": "work_commute", "label": "职场通勤"},
                {"value": "romantic_date", "label": "浪漫约会"},
                {"value": "smart_casual_gathering", "label": "轻正式聚会"},
                {"value": "holiday_gifting", "label": "节日礼赠"},
                {"value": "self_reward", "label": "自用犒赏"},
            ],
        )
        self.assertEqual(
            body["festivals"],
            ["情人节", "女神节", "520", "暑假", "开学季", "圣诞节"],
        )

    def test_selling_point_excel_upload_returns_catalog_and_entries(self):
        service = AiCopyService(FakeChatProvider(), FakeProductTool())
        router = create_ai_copy_router(service)
        endpoint = next(
            route.endpoint
            for route in router.routes
            if route.path == "/api/ai-copy/selling-point-catalog"
        )
        upload = UploadFile(
            filename="points.xlsx",
            file=BytesIO(
                build_selling_point_workbook(
                    [("SKU-001", "轻量透气，适合日常通勤")]
                )
            ),
        )

        result = asyncio.run(endpoint(upload))

        self.assertEqual(result.row_count, 1)
        self.assertEqual(result.entries[0].identifier, "SKU-001")
        self.assertTrue(result.catalog_id)

    def test_product_lookup_errors_are_mapped_without_leaking_request_key(self):
        class FailingTool(FakeProductTool):
            def inspect(self, _url, _config):
                raise ProductLookupError("商品服务暂时不可用")

        service = AiCopyService(FakeChatProvider(), FailingTool())
        router = create_ai_copy_router(service)
        endpoint = next(
            route.endpoint for route in router.routes if route.path.endswith("/product-references")
        )

        with self.assertRaises(HTTPException) as context:
            asyncio.run(
                endpoint(
                    ProductReferencesRequest(
                        product_urls=["https://shop.example/item/1"],
                        search={"api_key": None},
                    )
                )
            )

        self.assertEqual(context.exception.status_code, 502)
        self.assertEqual(
            context.exception.detail,
            "第 1 个商品链接读取失败：商品服务暂时不可用",
        )
