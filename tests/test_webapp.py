from __future__ import annotations

import asyncio
import json
import shutil
import tempfile
import threading
import time
import unittest
from dataclasses import replace
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock, patch

from fastapi import HTTPException, UploadFile
from loguru import logger
from openpyxl import Workbook, load_workbook

from uploader.errors import PublishResultUncertainError
from uploader.jd_video_uploader.main import JDVideo
from uploader.tmall_article_uploader.main import TmallArticle
from uploader.tmall_video_uploader.main import (
    TmallVideo,
    _contains_exact_product_id,
    _select_cover_ratio_and_continue,
    _has_explicit_empty_product_result,
    _normalize_option_text,
    _normalized_goods_ids,
    _two_character_chunks,
)
from utils.files import validate_media_filename
from webapp.api.batch import BatchValidationError
from webapp.api.batch_douyin_article import parse_douyin_article_batch_workbook
from webapp.api.batch_douyin_video import parse_douyin_video_batch_workbook
from webapp.api.batch_jd_video import parse_jd_video_batch_workbook
from webapp.api.batch_tmall_article import parse_tmall_article_batch_workbook
from webapp.api.batch_tmall_video import parse_tmall_video_batch_workbook
from webapp.api.batch_templates import build_batch_template
from webapp.api.batch_xiaohongshu_article import parse_xiaohongshu_article_batch_workbook
from webapp.api.batch_xiaohongshu_video import parse_xiaohongshu_video_batch_workbook
from webapp.api.agent_tasks import AgentTaskManager
from webapp.api.main import WebSettings
from webapp.api.batch import resolve_local_path
from webapp.api.main import create_app as _create_app
from webapp.api.models import ValidationError, validate_publish_request
from webapp.api.platforms import (
    JdVideoUploadRequest,
    TmallVideoUploadRequest,
    delete_account_cookie,
    resolve_account_file,
    secure_account_file,
    upload_jd_video,
    upload_tmall_video,
)
from webapp.api.store import JobStore
from webapp.api.tasks import TaskManager as _TaskManager
from webapp.auth import AuthService, AuthStore
from webapp.llm_adapter import LLMAdapterRegistry
from webapp.workspaces import AppDataPaths

TEST_USER_ID = "0" * 32


def TaskManager(store: JobStore, **kwargs) -> _TaskManager:
    """Build production TaskManager instances with isolated test user paths."""
    users_root = store.data_dir.parent / f".{store.data_dir.name}-test-users"
    paths = AppDataPaths.create(users_root).for_user(TEST_USER_ID)
    managed_upload_root = kwargs.pop("managed_upload_root", None)
    job_log_dir = kwargs.pop("job_log_dir", None)
    if managed_upload_root is not None:
        managed_upload_root.mkdir(parents=True, exist_ok=True)
        managed_upload_root.chmod(0o700)
        paths = replace(paths, uploads=managed_upload_root)
    if job_log_dir is not None:
        job_log_dir.mkdir(parents=True, exist_ok=True)
        job_log_dir.chmod(0o700)
        paths = replace(paths, job_logs=job_log_dir)
    return _TaskManager(store, user_id=TEST_USER_ID, paths=paths, **kwargs)


class _StaticWorkspaceRegistry:
    """Expose one fully assembled workspace to direct endpoint unit tests."""

    def __init__(self, workspace, settings: WebSettings) -> None:
        self.workspace = workspace
        self.ready = True
        self.user_workers = settings.user_workers
        self.global_browser_tasks = settings.global_browser_tasks

    def get(self, _user_id: str):
        return self.workspace

    def maintenance_errors(self) -> list[str]:
        return [
            f"{self.workspace.user_id}: {error}"
            for error in self.workspace.task_manager.maintenance_errors
        ]

    def delete_user_data(self, _user_id: str) -> None:
        self.workspace.task_manager.shutdown()
        shutil.rmtree(self.workspace.paths.root, ignore_errors=False)

    def close(self) -> None:
        return None


def create_app(
    settings: WebSettings,
    manager: _TaskManager,
):
    """Create an initialized app around an explicitly controlled test manager."""
    workspace = SimpleNamespace(
        user_id=TEST_USER_ID,
        paths=manager.paths,
        store=manager.store,
        task_manager=manager,
        llm_registry=LLMAdapterRegistry(),
        ai_copy_service=None,
    )
    registry = _StaticWorkspaceRegistry(workspace, settings)
    data_paths = AppDataPaths.create(settings.data_dir)
    auth_service = AuthService(AuthStore(data_paths.auth_database))
    auth_service.bootstrap_admin(
        username="testadmin",
        display_name="Test Admin",
        password="test-password-123",
    )
    app = _create_app(settings, registry, auth_service)
    app.state.test_workspace = workspace
    return app


class PublishRequestValidationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.video = Path(self.temp_dir.name) / "demo.mp4"
        self.video.write_bytes(b"video")

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_tmall_request_normalizes_tags(self):
        request = validate_publish_request(
            platform="tmall",
            cover_ratio="original",
            account="shop_1",
            video_path=self.video,
            original_filename="demo.mp4",
            title="夏季女鞋测评",
            description="轻便好穿",
            raw_tags="#女鞋, 夏季穿搭",
        )

        self.assertEqual(request.tags, ("女鞋", "夏季穿搭"))
        self.assertFalse(request.dry_run)

    def test_tmall_request_preserves_selected_cover_ratio(self):
        cover = Path(self.temp_dir.name) / "cover.png"
        cover.write_bytes(b"image")
        request = validate_publish_request(
            platform="tmall",
            cover_ratio="1:1",
            account="shop_1",
            video_path=self.video,
            cover_image_path=cover,
            original_filename="demo.mp4",
            title="夏季女鞋测评",
        )

        self.assertEqual(request.cover_ratio, "1:1")

    def test_tmall_rejects_unknown_cover_ratio(self):
        cover = Path(self.temp_dir.name) / "cover.png"
        cover.write_bytes(b"image")
        with self.assertRaisesRegex(ValidationError, "封面比例"):
            validate_publish_request(
                platform="tmall",
                cover_ratio="16:9",
                account="shop_1",
                video_path=self.video,
                cover_image_path=cover,
                original_filename="demo.mp4",
                title="夏季女鞋测评",
            )

    def test_tmall_video_without_custom_cover_rejects_crop_ratio(self):
        with self.assertRaisesRegex(ValidationError, "未上传自定义封面"):
            validate_publish_request(
                platform="tmall",
                account="shop_1",
                video_path=self.video,
                cover_ratio="1:1",
                original_filename="demo.mp4",
                title="夏季女鞋测评",
            )

    def test_tmall_custom_cover_requires_cover_ratio(self):
        cover = Path(self.temp_dir.name) / "cover.png"
        cover.write_bytes(b"image")
        with self.assertRaisesRegex(ValidationError, "封面比例"):
            validate_publish_request(
                platform="tmall",
                account="shop_1",
                video_path=self.video,
                cover_image_path=cover,
                original_filename="demo.mp4",
                title="夏季女鞋测评",
            )

    def test_parse_tags_accepts_chinese_commas(self):
        request = validate_publish_request(
            platform="tmall",
            cover_ratio="original",
            account="shop_1",
            video_path=self.video,
            original_filename="demo.mp4",
            title="夏季女鞋测评",
            description="轻便好穿",
            raw_tags="#女鞋， 夏季穿搭,通勤鞋",
        )

        self.assertEqual(request.tags, ("女鞋", "夏季穿搭", "通勤鞋"))

    def test_batch_path_parser_accepts_windows_absolute_path_on_non_windows_server(self):
        path = resolve_local_path(
            r"C:\Users\operator\Videos\demo.mp4", "视频路径"
        )
        self.assertEqual(str(path), r"C:\Users\operator\Videos\demo.mp4")

    def test_batch_path_parser_accepts_unc_windows_path(self):
        path = resolve_local_path(
            r"\\DESKTOP-01\素材\demo.mp4", "视频路径"
        )
        self.assertEqual(str(path), r"\\DESKTOP-01\素材\demo.mp4")

    def test_tmall_and_jd_accept_optional_cover_image(self):
        cover = Path(self.temp_dir.name) / "cover.png"
        cover.write_bytes(b"image")

        request = validate_publish_request(
            platform="tmall",
            cover_ratio="3:4",
            account="shop1",
            video_path=self.video,
            cover_image_path=cover,
            original_filename="demo.mp4",
            title="夏季女鞋测评",
        )

        self.assertEqual(request.cover_image_path, cover.resolve())

        jd_request = validate_publish_request(
            platform="jd",
            account="shop1",
            video_path=self.video,
            cover_image_path=cover,
            original_filename="demo.mp4",
            title="京东视频标题示例",
        )
        self.assertEqual(jd_request.cover_image_path, cover.resolve())

    def test_tmall_rejects_unsupported_cover_image(self):
        cover = Path(self.temp_dir.name) / "cover.gif"
        cover.write_bytes(b"image")

        with self.assertRaisesRegex(ValidationError, "JPG、PNG 或 WebP"):
            validate_publish_request(
                platform="tmall",
                cover_ratio="3:4",
                account="shop1",
                video_path=self.video,
                cover_image_path=cover,
                original_filename="demo.mp4",
                title="夏季女鞋测评",
            )

    def test_jd_rejects_description_and_invalid_title_length(self):
        with self.assertRaisesRegex(ValidationError, "独立文案"):
            validate_publish_request(
                platform="jd",
                account="shop1",
                video_path=self.video,
                original_filename="demo.mp4",
                title="京东视频标题",
                description="不支持",
            )

        with self.assertRaisesRegex(ValidationError, "5-27"):
            validate_publish_request(
                platform="jd",
                account="shop1",
                video_path=self.video,
                original_filename="demo.mp4",
                title="太短",
            )

    def test_xiaohongshu_and_douyin_accept_social_content(self):
        xhs_request = validate_publish_request(
            platform="xiaohongshu",
            account="shop1",
            video_path=self.video,
            original_filename="demo.mp4",
            title="小红书标题",
            description="正文内容",
            raw_tags="种草,穿搭",
        )
        self.assertEqual(xhs_request.tags, ("种草", "穿搭"))
        self.assertEqual(xhs_request.creator_declaration, "")

        douyin_request = validate_publish_request(
            platform="douyin",
            account="shop1",
            video_path=self.video,
            original_filename="demo.mp4",
            title="抖音标题",
            description="视频描述",
            raw_tags="热点,穿搭",
        )
        self.assertEqual(douyin_request.tags, ("热点", "穿搭"))
        self.assertEqual(douyin_request.creator_declaration, "")

    def test_xiaohongshu_and_douyin_reject_excess_tags(self):
        with self.assertRaisesRegex(ValidationError, "最多支持 20 个标签"):
            validate_publish_request(
                platform="xiaohongshu",
                account="shop1",
                video_path=self.video,
                original_filename="demo.mp4",
                title="小红书标题",
                raw_tags=",".join(f"标签{i}" for i in range(21)),
            )
        with self.assertRaisesRegex(ValidationError, "最多支持 20 个标签"):
            validate_publish_request(
                platform="douyin",
                account="shop1",
                video_path=self.video,
                original_filename="demo.mp4",
                title="抖音标题",
                raw_tags=",".join(f"标签{i}" for i in range(21)),
            )

    def test_tmall_rejects_combined_description_and_tag_overflow(self):
        with self.assertRaisesRegex(ValidationError, "文案与标签合计"):
            validate_publish_request(
                platform="tmall",
                cover_ratio="original",
                account="shop1",
                video_path=self.video,
                original_filename="demo.mp4",
                title="夏季女鞋测评",
                description="文" * 997,
                raw_tags="标签",
            )

    def test_empty_video_is_rejected_before_browser_startup(self):
        self.video.write_bytes(b"")

        with self.assertRaisesRegex(ValidationError, "视频文件为空"):
            validate_publish_request(
                platform="tmall",
                cover_ratio="original",
                account="shop1",
                video_path=self.video,
                original_filename="demo.mp4",
                title="夏季女鞋测评",
            )

    def test_creator_declaration_is_validated(self):
        with self.assertRaisesRegex(ValidationError, "创作者声明"):
            validate_publish_request(
                platform="tmall",
                cover_ratio="original",
                account="shop1",
                video_path=self.video,
                original_filename="demo.mp4",
                title="夏季女鞋测评",
                raw_creator_declaration="随便填写",
            )

    def test_schedule_requires_two_hour_lead_time(self):
        near_future = (datetime.now() + timedelta(minutes=30)).strftime("%Y-%m-%d %H:%M")
        with self.assertRaisesRegex(ValidationError, "2 小时"):
            validate_publish_request(
                platform="jd",
                account="shop1",
                video_path=self.video,
                original_filename="demo.mp4",
                title="京东视频标题示例",
                raw_schedule=near_future,
            )

    def test_tmall_product_id_matching_uses_numeric_boundaries(self):
        self.assertTrue(_contains_exact_product_id('data-item-id="123"', "123"))
        self.assertFalse(_contains_exact_product_id('data-item-id="1234"', "123"))
        self.assertFalse(_contains_exact_product_id('data-item-id="9123"', "123"))

    def test_video_filename_is_portable_to_windows_agent(self):
        with self.assertRaisesRegex(ValueError, "Windows 保留名称"):
            validate_media_filename("CON.mp4")
        self.assertEqual(validate_media_filename("campaign.mp4"), "campaign.mp4")

    def test_tmall_accepts_up_to_six_unique_product_ids(self):
        request = validate_publish_request(
            platform="tmall",
            cover_ratio="original",
            account="shop1",
            video_path=self.video,
            original_filename="demo.mp4",
            title="夏季女鞋测评",
            goods_id="123，456\n123 789",
        )

        self.assertEqual(request.goods_id, "123,456,789")
        self.assertEqual(_normalized_goods_ids(request.goods_id), ("123", "456", "789"))

    def test_tmall_rejects_more_than_six_product_ids(self):
        with self.assertRaisesRegex(ValidationError, "最多关联 6 个"):
            validate_publish_request(
                platform="tmall",
                cover_ratio="original",
                account="shop1",
                video_path=self.video,
                original_filename="demo.mp4",
                title="夏季女鞋测评",
                goods_id="1,2,3,4,5,6,7",
            )

    def test_jd_accepts_ten_product_ids_and_rejects_more(self):
        request = validate_publish_request(
            platform="jd",
            account="shop1",
            video_path=self.video,
            original_filename="demo.mp4",
            title="京东视频标题示例",
            goods_id="1,2,3,4,5,6,7,8,9,10",
        )
        self.assertEqual(request.goods_id, "1,2,3,4,5,6,7,8,9,10")

        with self.assertRaisesRegex(ValidationError, "最多关联 10 个"):
            validate_publish_request(
                platform="jd",
                account="shop1",
                video_path=self.video,
                original_filename="demo.mp4",
                title="京东视频标题示例",
                goods_id="1,2,3,4,5,6,7,8,9,10,11",
            )

    def test_jd_article_accepts_images_body_and_topic(self):
        images = []
        for index in range(2):
            image = Path(self.temp_dir.name) / f"article-{index}.jpg"
            image.write_bytes(b"image")
            images.append(image)

        request = validate_publish_request(
            platform="jd",
            account="shop1",
            content_type="article",
            image_paths=tuple(images),
            original_filename=images[0].name,
            title="京东图文标题",
            description="京东图文正文",
            activity_topic="数码先锋",
        )

        self.assertEqual(request.image_paths, tuple(path.resolve() for path in images))
        self.assertEqual(request.description, "京东图文正文")
        self.assertEqual(request.activity_topic, "数码先锋")

    def test_tmall_end_of_list_is_not_treated_as_an_empty_result(self):
        self.assertFalse(_has_explicit_empty_product_result("¥149.00\n没有更多了"))
        self.assertTrue(_has_explicit_empty_product_result("暂无商品"))

    def test_tmall_creator_declaration_ignores_dom_whitespace(self):
        self.assertEqual(_normalize_option_text(" 内容 无需\n标注 "), "内容无需标注")

    def test_tmall_music_name_is_normalized_and_jd_rejects_it(self):
        request = validate_publish_request(
            platform="tmall",
            cover_ratio="original",
            account="shop1",
            video_path=self.video,
            original_filename="demo.mp4",
            title="夏季女鞋测评",
            raw_music_name=" 默契 ",
        )
        self.assertEqual(request.music_name, "默契")

        with self.assertRaisesRegex(ValidationError, "音乐字段"):
            validate_publish_request(
                platform="jd",
                account="shop1",
                video_path=self.video,
                original_filename="demo.mp4",
                title="京东视频标题示例",
                raw_music_name="默契",
            )

    def test_tmall_music_search_uses_two_character_input_chunks(self):
        self.assertEqual(_two_character_chunks("默契"), ("默契",))
        self.assertEqual(_two_character_chunks("夏日好物"), ("夏日", "好物"))
        self.assertEqual(_two_character_chunks("abcde"), ("ab", "cd", "e"))

    def test_tmall_custom_cover_uses_the_current_two_dialog_flow(self):
        cover = Path(self.temp_dir.name) / "20260811-093942.jpeg"
        cover.write_bytes(b"cover")
        uploader = object.__new__(TmallVideo)
        uploader.cover_image_path = str(cover)
        uploader.cover_ratio = "3:4"

        def locator():
            item = MagicMock()
            item.wait_for = AsyncMock()
            item.click = AsyncMock()
            return item

        def query(item):
            result = MagicMock()
            result.first = item
            result.last = item
            return result

        edit = locator()
        smart_cover_loading = locator()
        cover_upload = locator()
        cover_dialog = locator()
        cover_dialog.get_by_text.return_value = query(cover_upload)
        opened_overlays = MagicMock()
        opened_overlays.count = AsyncMock(side_effect=[0, 0, 1])
        opened_overlays.nth.return_value = cover_dialog

        frame = MagicMock()
        frame.locator.side_effect = lambda selector: (
            query(edit)
            if selector == '[data-autolog-container="coverOperate_edit"]'
            else opened_overlays
        )
        frame.get_by_text.return_value = query(smart_cover_loading)

        page = MagicMock()
        picker_frame = MagicMock()
        picker_frame.url = "https://market.m.taobao.com/app/crs-qn/sucai-selector-ng/index"
        picker_frame.evaluate = AsyncMock(
            return_value={"count": 1, "checked": False, "matchingCards": ["new cover"]}
        )
        selected_control = MagicMock()
        selected_control.evaluate = AsyncMock()
        selected_control.is_checked = AsyncMock(return_value=True)
        picker_frame.locator.return_value = selected_control
        page.frames = [picker_frame]

        upload_picker_file = AsyncMock()
        click_visible_frame_button = AsyncMock()
        select_cover_ratio_and_continue = AsyncMock()

        with (
            patch(
                "uploader.tmall_video_uploader.main.asyncio.sleep",
                new=AsyncMock(),
            ),
            patch(
                "uploader.tmall_video_uploader.main._upload_picker_file",
                new=upload_picker_file,
            ),
            patch(
                "uploader.tmall_video_uploader.main._cover_picker_upload_name",
                return_value="mpau-cover-20260817-091530-a1b2c3d4e5f6.jpeg",
            ),
            patch(
                "uploader.tmall_video_uploader.main._click_visible_frame_button",
                new=click_visible_frame_button,
            ),
            patch(
                "uploader.tmall_video_uploader.main._select_cover_ratio_and_continue",
                new=select_cover_ratio_and_continue,
            ),
        ):
            asyncio.run(uploader._set_custom_cover(frame, page))

        edit.click.assert_awaited_once_with(timeout=10000)
        cover_upload.click.assert_awaited_once_with()
        upload_picker_file.assert_awaited_once()
        uploaded_cover_path = upload_picker_file.await_args.args[2]
        self.assertEqual(uploaded_cover_path.name, "mpau-cover-20260817-091530-a1b2c3d4e5f6.jpeg")
        select_cover_ratio_and_continue.assert_awaited_once_with(page, "3:4")
        selection_script, selection_stem = picker_frame.evaluate.await_args.args
        self.assertIn("expectedStem", selection_script)
        self.assertEqual(selection_stem, "mpau-cover-20260817-091530-a1b2c3d4e5f6")
        self.assertNotIn("document.images", selection_script)
        selected_control.evaluate.assert_awaited_once_with("(control) => control.click()")
        self.assertEqual(
            [call.args for call in click_visible_frame_button.await_args_list],
            [
                ((picker_frame,), ("完成",)),
                ((picker_frame,), ("确定",)),
                ((picker_frame,), ("下一步", "完成", "确定")),
            ],
        )
        frame.get_by_text.assert_called_once_with("智能封面图生成中", exact=False)

    def test_tmall_cover_crop_chooses_requested_one_to_one_card(self):
        from PIL import Image

        image = Image.new("RGB", (1280, 900), "white")
        buffer = BytesIO()
        image.save(buffer, "PNG")
        cards = (
            (600.0, 532.0, 687.0, 532.0),
            (710.0, 532.0, 797.0, 532.0),
            (820.0, 532.0, 907.0, 532.0),
        )
        page = MagicMock()
        page.screenshot = AsyncMock(return_value=buffer.getvalue())
        page.evaluate = AsyncMock(return_value={"width": 1280, "height": 900})
        page.mouse.click = AsyncMock()

        with (
            patch("uploader.tmall_video_uploader.main._find_cover_ratio_cards", return_value=cards),
            patch("uploader.tmall_video_uploader.main._ratio_card_is_selected", return_value=True),
            patch("uploader.tmall_video_uploader.main._find_cover_next_button", return_value=(1030.0, 824.5)),
            patch("uploader.tmall_video_uploader.main.asyncio.sleep", new=AsyncMock()),
        ):
            asyncio.run(_select_cover_ratio_and_continue(page, "1:1"))

        self.assertEqual(page.mouse.click.await_args_list[0].args, ((820.0 + 907.0) / 2, 532.0))

    def test_tmall_original_cover_ratio_skips_ratio_card_click(self):
        from PIL import Image

        image = Image.new("RGB", (1280, 900), "white")
        buffer = BytesIO()
        image.save(buffer, "PNG")
        page = MagicMock()
        page.screenshot = AsyncMock(return_value=buffer.getvalue())
        page.evaluate = AsyncMock(return_value={"width": 1280, "height": 900})
        page.mouse.click = AsyncMock()

        with patch(
            "uploader.tmall_video_uploader.main._find_cover_next_button",
            return_value=(1030.0, 824.5),
        ):
            asyncio.run(
                _select_cover_ratio_and_continue(page, "original")
            )

        page.mouse.click.assert_awaited_once_with(1030.0, 824.5, delay=150)

    def test_tmall_article_original_ratio_skips_crop_flow(self):
        uploader = object.__new__(TmallArticle)
        uploader.cover_ratio = "original"
        uploader._crop_uploaded_images = AsyncMock()

        asyncio.run(uploader._crop_images_if_requested(MagicMock()))

        uploader._crop_uploaded_images.assert_not_awaited()

    def test_tmall_article_one_to_one_ratio_triggers_requested_crop(self):
        uploader = object.__new__(TmallArticle)
        uploader.cover_ratio = "1:1"
        uploader._crop_uploaded_images = AsyncMock()
        frame = MagicMock()

        asyncio.run(uploader._crop_images_if_requested(frame))

        uploader._crop_uploaded_images.assert_awaited_once_with(frame, "1:1")

    def test_unknown_tmall_navigation_is_not_publish_confirmation(self):
        class Body:
            async def inner_text(self, **_kwargs):
                return "页面处理中"

        class Surface:
            url = "https://creator.guanghe.taobao.com/page/unknown"

            def locator(self, _selector):
                return Body()

        uploader = object.__new__(TmallVideo)
        with patch("uploader.tmall_video_uploader.main.asyncio.sleep", new=AsyncMock()):
            with self.assertRaises(PublishResultUncertainError):
                asyncio.run(
                    uploader._wait_for_publish_confirmation(
                        Surface(),
                        Surface(),
                        initial_url=Surface.url,
                        before_text="",
                        timeout_seconds=1,
                    )
                )

    def test_jd_captcha_waiting_supports_windowed_process_without_stdin(self):
        uploader = object.__new__(JDVideo)
        frame = SimpleNamespace(evaluate=AsyncMock(side_effect=[True, False]))
        with (
            patch("uploader.jd_video_uploader.main.sys.stdin", SimpleNamespace(isatty=lambda: False)),
            patch("uploader.jd_video_uploader.main.asyncio.sleep", new=AsyncMock()),
        ):
            asyncio.run(uploader._handle_captcha(frame))

    def test_jd_wait_for_video_uploaded_recovers_from_iframe_reload(self):
        uploader = object.__new__(JDVideo)

        def make_frame(*, detached: bool = False):
            body = MagicMock()
            if detached:
                body.inner_text = AsyncMock(side_effect=RuntimeError("Frame was detached"))
            else:
                body.inner_text = AsyncMock(return_value="等待视频上传")

            edit_button = MagicMock()
            edit_button.count = AsyncMock(return_value=1)
            edit_button.is_visible = AsyncMock(return_value=True)

            edit_locator = MagicMock()
            edit_locator.filter.return_value.first = edit_button

            frame = MagicMock()

            def locator(selector):
                if selector == "body":
                    return body
                if selector == ".edit-cover-btn":
                    return edit_locator
                raise AssertionError(f"unexpected selector: {selector}")

            frame.locator.side_effect = locator
            return frame

        first_frame = make_frame(detached=True)
        second_frame = make_frame()
        page = SimpleNamespace(url="https://dr.jd.com/jm/#/n/publish-video.html?platform=jm-pop", is_closed=lambda: False)

        with (
            patch("uploader.jd_video_uploader.main._find_publish_iframe", new=AsyncMock(return_value=second_frame)),
            patch("uploader.jd_video_uploader.main.asyncio.sleep", new=AsyncMock()),
        ):
            result = asyncio.run(uploader._wait_for_video_uploaded(page, first_frame, timeout_seconds=5))

        self.assertIs(result, second_frame)

    def test_jd_set_custom_cover_recovers_from_iframe_reload(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            cover = Path(temp_dir) / "cover.jpg"
            cover.write_bytes(b"image")
            uploader = object.__new__(JDVideo)
            uploader.cover_image_path = str(cover)

            def make_frame(*, detached: bool = False):
                def first_locator(target):
                    outer = MagicMock()
                    outer.first = target
                    return outer

                edit_button = MagicMock()
                edit_button.count = AsyncMock(return_value=1)
                edit_button.wait_for = AsyncMock(
                    side_effect=RuntimeError("Frame was detached") if detached else None
                )
                edit_button.click = AsyncMock()

                preview = MagicMock()
                preview.count = AsyncMock(return_value=0)

                file_input = MagicMock()
                file_input.count = AsyncMock(return_value=1)
                file_input.set_input_files = AsyncMock()

                confirm_button = MagicMock()
                confirm_button.wait_for = AsyncMock()
                confirm_button.click = AsyncMock()

                modal = MagicMock()
                modal.wait_for = AsyncMock()
                crop_preview = MagicMock()
                crop_preview.count = AsyncMock(return_value=0)
                modal.locator.return_value.last = crop_preview

                frame = MagicMock()

                def locator(selector):
                    if selector in {
                        '[data-spm-click="openVideoCoverModal"]',
                        ".edit-cover-btn",
                    }:
                        return first_locator(edit_button)
                    if selector == ".video-cover-wrapper .preview-img":
                        return first_locator(preview)
                    if selector == 'input[type="file"][accept*="image"]':
                        return first_locator(file_input)
                    if selector == ".jd-modal-wrap":
                        outer = MagicMock()
                        outer.last = modal
                        return outer
                    if selector == 'button[data-component-label="确定"]':
                        return first_locator(confirm_button)
                    raise AssertionError(f"unexpected selector: {selector}")

                frame.locator.side_effect = locator
                return frame

            first_frame = make_frame(detached=True)
            second_frame = make_frame()
            page = SimpleNamespace(url="https://dr.jd.com/jm/#/n/publish-video.html?platform=jm-pop", is_closed=lambda: False)

            with (
                patch("uploader.jd_video_uploader.main._find_publish_iframe", new=AsyncMock(return_value=second_frame)),
                patch("uploader.jd_video_uploader.main.asyncio.sleep", new=AsyncMock()),
            ):
                result = asyncio.run(uploader._set_custom_cover(page, first_frame))

            self.assertIs(result, second_frame)


class TaskManagerTests(unittest.TestCase):
    @staticmethod
    def wait_for_status(store: JobStore, job_id: str, status: str) -> dict:
        for _ in range(100):
            job = store.get_job(job_id)
            if job and job["status"] == status:
                return job
            time.sleep(0.02)
        raise AssertionError(f"job {job_id} did not reach {status}")

    def test_account_tasks_are_persisted_and_completed(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            manager = TaskManager(
                store,
                runner=lambda job: {"message": f"{job['kind']} complete"},
                max_workers=1,
            )
            try:
                job = manager.submit_account_task(
                    kind="check", platform="tmall", account="shop1", headed=False
                )
                for _ in range(50):
                    completed = store.get_job(job["id"])
                    if completed and completed["status"] == "succeeded":
                        break
                    time.sleep(0.02)
                else:
                    self.fail("background task did not complete")

                self.assertEqual(completed["message"], "check complete")
                self.assertEqual(store.list_accounts()[0]["account"], "shop1")
            finally:
                manager.shutdown()

    def test_batch_publish_task_keeps_excel_row_metadata(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            video = Path(temp_dir) / "demo.mp4"
            video.write_bytes(b"video")
            request = validate_publish_request(
                platform="tmall",
                cover_ratio="original",
                account="shop1",
                video_path=video,
                original_filename=video.name,
                title="夏季女鞋测评",
            )
            store = JobStore(Path(temp_dir) / "state")
            manager = TaskManager(store, runner=lambda job: {"message": "complete"}, max_workers=1)
            try:
                job = manager.submit_publish_task(request, batch_id="batch-1", source_row=5)
                for _ in range(50):
                    completed = store.get_job(job["id"])
                    if completed and completed["status"] == "succeeded":
                        break
                    time.sleep(0.02)
                else:
                    self.fail("batch publish task did not complete")

                self.assertEqual(completed["batch_id"], "batch-1")
                self.assertEqual(completed["source_row"], 5)
            finally:
                manager.shutdown()

    def test_running_browser_task_can_be_cancelled(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            manager = TaskManager(store, max_workers=1)
            started = threading.Event()

            async def wait_for_cancellation(_job):
                started.set()
                await asyncio.Event().wait()

            manager._run_platform_task_async = wait_for_cancellation
            try:
                job = manager.submit_account_task(
                    kind="login", platform="tmall", account="shop1", headed=True
                )
                self.assertTrue(started.wait(timeout=1))

                cancelling = manager.cancel_task(job["id"])
                self.assertEqual(cancelling["status"], "cancelling")
                for _ in range(50):
                    cancelled = store.get_job(job["id"])
                    if cancelled and cancelled["status"] == "cancelled":
                        break
                    time.sleep(0.02)
                else:
                    self.fail("browser task did not cancel")

                self.assertEqual(cancelled["message"], "浏览器任务已中断")
            finally:
                manager.shutdown()

    def test_uncertain_publish_result_has_a_distinct_terminal_status(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            media_dir = AppDataPaths.create(
                root / ".runtime-test-users"
            ).for_user(TEST_USER_ID).media
            video = media_dir / "demo.mp4"
            video.write_bytes(b"video")
            request = validate_publish_request(
                platform="tmall",
                cover_ratio="original",
                account="shop1",
                video_path=video,
                original_filename=video.name,
                title="夏季女鞋测评",
            )

            def uncertain_runner(_job):
                raise PublishResultUncertainError("发布按钮已点击")

            store = JobStore(root / "state")
            manager = TaskManager(store, runner=uncertain_runner, max_workers=1)
            try:
                job = manager.submit_publish_task(request)
                completed = self.wait_for_status(store, job["id"], "uncertain")
                self.assertIn("不要重试", completed["message"])
                self.assertIn("发布按钮已点击", completed["error"])
            finally:
                manager.shutdown()

    def test_late_cancel_does_not_overwrite_a_completed_result(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            started = threading.Event()
            release = threading.Event()

            def runner(_job):
                started.set()
                release.wait(timeout=2)
                return {"message": "platform confirmed"}

            manager = TaskManager(store, runner=runner, max_workers=1)
            try:
                job = manager.submit_account_task(
                    kind="check", platform="tmall", account="shop1", headed=False
                )
                self.assertTrue(started.wait(timeout=1))
                manager.cancel_task(job["id"])
                release.set()
                completed = self.wait_for_status(store, job["id"], "succeeded")
                self.assertEqual(completed["message"], "platform confirmed")
            finally:
                release.set()
                manager.shutdown()

    def test_same_account_is_fifo_without_blocking_another_account(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            first_started = threading.Event()
            release_first = threading.Event()
            other_finished = threading.Event()
            execution_order: list[str] = []

            def runner(job):
                execution_order.append(job["kind"])
                if job["kind"] == "first":
                    first_started.set()
                    release_first.wait(timeout=2)
                if job["kind"] == "other":
                    other_finished.set()
                return {"message": "complete"}

            manager = TaskManager(store, runner=runner, max_workers=2)
            try:
                first = manager.submit_account_task(
                    kind="first", platform="tmall", account="shop1"
                )
                self.assertTrue(first_started.wait(timeout=1))
                second = manager.submit_account_task(
                    kind="second", platform="tmall", account="shop1"
                )
                third = manager.submit_account_task(
                    kind="third", platform="tmall", account="shop1"
                )
                other = manager.submit_account_task(
                    kind="other", platform="tmall", account="shop2"
                )

                self.assertTrue(other_finished.wait(timeout=1))
                self.assertEqual(store.get_job(second["id"])["status"], "queued")
                release_first.set()
                for job in (first, second, third, other):
                    self.wait_for_status(store, job["id"], "succeeded")
                self.assertLess(execution_order.index("second"), execution_order.index("third"))
            finally:
                release_first.set()
                manager.shutdown()

    def test_startup_recovers_queued_and_terminalizes_interrupted_jobs(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            queued = store.create_job(
                kind="check", platform="tmall", account="shop1", payload={"headed": False}
            )
            running = store.create_job(
                kind="check", platform="tmall", account="shop2", payload={"headed": False}
            )
            cancelling = store.create_job(
                kind="check", platform="jd", account="shop3", payload={"headed": False}
            )
            store.update_job(running["id"], status="running")
            store.update_job(cancelling["id"], status="cancelling")

            manager = TaskManager(store, runner=lambda _job: {"message": "recovered"})
            try:
                self.assertEqual(store.get_job(queued["id"])["status"], "queued")
                self.assertEqual(store.get_job(running["id"])["status"], "running")
                manager.start()
                self.wait_for_status(store, queued["id"], "succeeded")
                self.assertEqual(store.get_job(running["id"])["status"], "failed")
                self.assertIn("结果可能不确定", store.get_job(running["id"])["message"])
                self.assertEqual(store.get_job(cancelling["id"])["status"], "cancelled")
            finally:
                manager.shutdown()

    def test_semantically_invalid_state_is_quarantined_on_startup(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / "state.json").write_text(
                '{"accounts":[{}],"jobs":{"bad":null}}', encoding="utf-8"
            )

            store = JobStore(root)

            self.assertEqual(store.list_accounts(), [])
            self.assertEqual(store.list_jobs(), [])
            self.assertEqual(len(list(root.glob("state.json.corrupt-*"))), 1)

    def test_interrupted_publish_is_recovered_as_uncertain(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            job = store.create_job(
                kind="publish",
                platform="tmall",
                account="shop1",
                payload={"video_path": "/tmp/demo.mp4"},
            )
            store.update_job(job["id"], status="running")

            store.recover_interrupted_jobs()

            recovered = store.get_job(job["id"])
            self.assertEqual(recovered["status"], "uncertain")
            self.assertIn("结果可能不确定", recovered["message"])

    def test_second_manager_cannot_recover_a_live_runtime(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            started = threading.Event()
            release = threading.Event()

            def runner(_job):
                started.set()
                release.wait(timeout=2)
                return {"message": "complete"}

            first = TaskManager(store, runner=runner, max_workers=1)
            second = TaskManager(JobStore(Path(temp_dir)), runner=lambda _job: {"message": "wrong"})
            try:
                job = first.submit_account_task(
                    kind="check", platform="tmall", account="shop1", headed=False
                )
                self.assertTrue(started.wait(timeout=1))
                self.assertEqual(store.get_job(job["id"])["status"], "running")

                with self.assertRaisesRegex(RuntimeError, "已有任务管理进程"):
                    second.start()

                self.assertEqual(store.get_job(job["id"])["status"], "running")
                self.assertTrue(first.ready)
                self.assertFalse(second.ready)
                release.set()
                self.wait_for_status(store, job["id"], "succeeded")
            finally:
                release.set()
                second.shutdown()
                first.shutdown()

    def test_startup_removes_orphans_but_preserves_a_referenced_upload(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            uploads = root / "uploads"
            orphan = uploads / ("a" * 32)
            referenced = uploads / ("b" * 32)
            orphan.mkdir(parents=True)
            referenced.mkdir()
            (orphan / "old.mp4").write_bytes(b"old")
            video = referenced / "queued.mp4"
            video.write_bytes(b"video")
            store = JobStore(root / "state")
            store.create_job(
                kind="publish",
                platform="tmall",
                account="shop1",
                payload={"managed_upload": True, "video_path": str(video)},
            )
            started = threading.Event()
            release = threading.Event()

            def runner(_job):
                started.set()
                release.wait(timeout=2)
                return {"message": "complete"}

            manager = TaskManager(store, runner=runner, managed_upload_root=uploads)
            try:
                manager.start()
                self.assertTrue(started.wait(timeout=1))
                self.assertFalse(orphan.exists())
                self.assertTrue(referenced.exists())
            finally:
                release.set()
                manager.wait_for_account_idle("tmall", "shop1", timeout=2)
                manager.shutdown()

    def test_managed_upload_cleanup_failure_is_recorded_on_the_job(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            uploads = root / "uploads"
            upload_dir = uploads / ("c" * 32)
            upload_dir.mkdir(parents=True)
            video = upload_dir / "demo.mp4"
            video.write_bytes(b"video")
            request = validate_publish_request(
                platform="tmall",
                cover_ratio="original",
                account="shop1",
                video_path=video,
                original_filename=video.name,
                title="夏季女鞋测评",
                managed_upload=True,
            )
            store = JobStore(root / "state")
            manager = TaskManager(
                store,
                runner=lambda _job: {"message": "complete"},
                managed_upload_root=uploads,
            )
            try:
                with patch.object(
                    manager, "_cleanup_managed_upload", side_effect=OSError("permission denied")
                ):
                    job = manager.submit_publish_task(request)
                    self.assertTrue(manager.wait_for_account_idle("tmall", "shop1", timeout=2))

                completed = store.get_job(job["id"])
                self.assertEqual(completed["status"], "succeeded")
                self.assertEqual(completed["result"]["cleanup_error"], "permission denied")
                self.assertIn("临时视频清理失败", completed["message"])
                self.assertTrue(manager.maintenance_errors)
            finally:
                manager.shutdown()

    def test_publish_preflight_rejects_expired_schedule_and_missing_video(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            manager = TaskManager(store)
            base_job = {
                "id": "preflight",
                "kind": "publish",
                "platform": "tmall",
                "account": "shop1",
                "payload": {
                    "headed": False,
                    "video_path": str(Path(temp_dir) / "missing.mp4"),
                },
            }
            try:
                expired_job = {
                    **base_job,
                    "payload": {
                        **base_job["payload"],
                        "schedule": "2000-01-01T00:00:00",
                    },
                }
                with self.assertRaisesRegex(RuntimeError, "定时发布时间已过"):
                    asyncio.run(manager._run_platform_task_async(expired_job))

                missing_video_job = {
                    **base_job,
                    "payload": {**base_job["payload"], "schedule": None},
                }
                with self.assertRaisesRegex(RuntimeError, "移动或删除"):
                    asyncio.run(manager._run_platform_task_async(missing_video_job))
            finally:
                manager.shutdown()

    def test_jd_task_maps_the_web_payload_to_the_adapter_contract(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            video = root / "demo.mp4"
            video.write_bytes(b"video")
            manager = TaskManager(JobStore(root / "state"))
            job = {
                "id": "jd-contract",
                "kind": "publish",
                "platform": "jd",
                "account": "shop1",
                "payload": {
                    "headed": True,
                    "video_path": str(video),
                    "title": "京东视频标题示例",
                    "goods_id": "12345",
                    "schedule": None,
                    "original": True,
                    "dry_run": True,
                },
            }
            upload = AsyncMock()
            session_pool = object()
            try:
                with patch.object(
                    manager.browser_runtime,
                    "is_current_loop",
                    return_value=True,
                ), patch.object(
                    manager.browser_runtime,
                    "jd_sessions",
                    return_value=session_pool,
                ), patch("webapp.api.platforms.upload_jd_video", new=upload):
                    result = asyncio.run(manager._run_platform_task_async(job))

                upload.assert_awaited_once()
                request = upload.await_args.args[0]
                self.assertIsInstance(request, JdVideoUploadRequest)
                self.assertEqual(request.account_name, "shop1")
                self.assertEqual(request.video_file, video)
                self.assertEqual(request.title, "京东视频标题示例")
                self.assertEqual(request.goods_id, "12345")
                self.assertTrue(request.original)
                self.assertFalse(request.headless)
                self.assertTrue(request.dry_run)
                self.assertIs(upload.await_args.kwargs["session_pool"], session_pool)
                self.assertEqual(result["message"], "流程验证已完成，未提交发布")
            finally:
                manager.shutdown()

    def test_managed_web_upload_is_removed_after_terminal_status(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            uploads = root / "uploads"
            upload_dir = uploads / ("a" * 32)
            upload_dir.mkdir(parents=True)
            video = upload_dir / "demo.mp4"
            video.write_bytes(b"video")
            request = validate_publish_request(
                platform="tmall",
                cover_ratio="original",
                account="shop1",
                video_path=video,
                original_filename=video.name,
                title="夏季女鞋测评",
                managed_upload=True,
            )
            store = JobStore(root / "state")
            manager = TaskManager(
                store,
                runner=lambda _job: {"message": "complete"},
                managed_upload_root=uploads,
            )
            try:
                job = manager.submit_publish_task(request)
                self.wait_for_status(store, job["id"], "succeeded")
                for _ in range(50):
                    if not upload_dir.exists():
                        break
                    time.sleep(0.02)
                self.assertFalse(upload_dir.exists())
            finally:
                manager.shutdown()

    def test_cancel_account_tasks_cancels_running_and_queued_work(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            manager = TaskManager(store, max_workers=1)
            started = threading.Event()

            async def wait_for_cancellation(_job):
                started.set()
                await asyncio.Event().wait()

            manager._run_platform_task_async = wait_for_cancellation
            try:
                jobs = [
                    manager.submit_account_task(
                        kind="login", platform="tmall", account="shop1", headed=True
                    )
                    for _ in range(3)
                ]
                self.assertTrue(started.wait(timeout=1))
                affected = manager.cancel_account_tasks("tmall", "shop1")
                self.assertEqual(len(affected), 3)
                self.assertTrue(manager.wait_for_account_idle("tmall", "shop1", timeout=2))
                self.assertTrue(
                    all(store.get_job(job["id"])["status"] == "cancelled" for job in jobs)
                )
            finally:
                manager.shutdown()

    def test_job_log_failure_marks_task_failed_instead_of_leaving_it_running(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir) / "state")
            manager = TaskManager(
                store,
                runner=lambda _job: {"message": "complete"},
                job_log_dir=Path(temp_dir) / "logs",
            )
            manager._attach_job_log = lambda _job: (_ for _ in ()).throw(
                OSError("log directory unavailable")
            )
            try:
                job = manager.submit_account_task(
                    kind="check", platform="tmall", account="shop1", headed=False
                )
                failed = self.wait_for_status(store, job["id"], "failed")
                self.assertIn("log directory unavailable", failed["error"])
            finally:
                manager.shutdown()

    def test_each_task_gets_an_independent_deletable_log(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            log_dir = Path(temp_dir) / "logs"
            store = JobStore(Path(temp_dir) / "state")

            def runner(_job):
                logger.bind(business_name="tmall").info("job-specific-entry")
                return {"message": "complete"}

            manager = TaskManager(store, runner=runner, job_log_dir=log_dir)
            try:
                job = manager.submit_account_task(
                    kind="check", platform="tmall", account="shop1", headed=False
                )
                self.wait_for_status(store, job["id"], "succeeded")
                self.assertTrue(manager.wait_for_account_idle("tmall", "shop1", timeout=1))
                log_path = manager.job_log_path(job["id"])
                self.assertIsNotNone(log_path)
                self.assertIn("job-specific-entry", log_path.read_text(encoding="utf-8"))
                self.assertEqual(log_dir.stat().st_mode & 0o777, 0o700)
                self.assertEqual(log_path.stat().st_mode & 0o777, 0o600)

                manager.delete_job_artifacts(job["id"])
                self.assertFalse(log_path.exists())
            finally:
                manager.shutdown()


class TmallBatchWorkbookTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.base_dir = Path(self.temp_dir.name)
        self.video = self.base_dir / "photos" / "demo.mp4"
        self.video.parent.mkdir()
        self.video.write_bytes(b"video")
        self.cover = self.base_dir / "photos" / "cover.png"
        self.cover.write_bytes(b"cover")

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def build_workbook(self, rows: list[list[object]]) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(
            [
                "视频路径",
                "自定义封面",
                "封面比例",
                "标题",
                "文案",
                "标签",
                "商品ID",
                "活动话题",
                "定时发布",
                "创作者声明",
            ]
        )
        for row in rows:
            worksheet.append([*row, "内容无需标注"])
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    def test_valid_rows_map_to_tmall_publish_requests(self):
        content = self.build_workbook(
            [[str(self.video), "", "", "夏季女鞋穿搭", "轻盈舒适", "女鞋，夏季穿搭", "12345，67890", "夏日上新", ""]]
        )

        rows = parse_tmall_video_batch_workbook(
            content,
            account="shop1",
            dry_run=True,
            headed=True,
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].row_number, 2)
        self.assertEqual(rows[0].request.video_path, self.video.resolve())
        self.assertEqual(rows[0].request.tags, ("女鞋", "夏季穿搭"))
        self.assertEqual(rows[0].request.goods_id, "12345,67890")
        self.assertEqual(rows[0].request.cover_ratio, "original")
        self.assertTrue(rows[0].request.dry_run)

    def test_video_without_cover_columns_uses_platform_generated_cover(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["视频路径", "标题", "创作者声明"])
        worksheet.append([str(self.video), "夏季女鞋穿搭", "内容无需标注"])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        rows = parse_tmall_video_batch_workbook(
            output.getvalue(), account="shop1", dry_run=True, headed=False
        )

        self.assertIsNone(rows[0].request.cover_image_path)
        self.assertEqual(rows[0].request.cover_ratio, "original")

    def test_video_rows_map_each_selected_cover_ratio(self):
        content = self.build_workbook(
            [
                [str(self.video), str(self.cover), "原始", "原始比例", "", "", "", "", ""],
                [str(self.video), str(self.cover), "3:4", "三比四比例", "", "", "", "", ""],
                [str(self.video), str(self.cover), "1:1", "一比一比例", "", "", "", "", ""],
            ]
        )

        rows = parse_tmall_video_batch_workbook(
            content, account="shop1", dry_run=True, headed=True
        )

        self.assertEqual(
            [row.request.cover_ratio for row in rows], ["original", "3:4", "1:1"]
        )

    def test_video_without_custom_cover_rejects_populated_ratio(self):
        content = self.build_workbook(
            [[str(self.video), "", "3:4", "夏季女鞋穿搭", "", "", "", "", ""]]
        )

        with self.assertRaises(BatchValidationError) as context:
            parse_tmall_video_batch_workbook(
                content, account="shop1", dry_run=True, headed=True
            )

        self.assertIn("封面比例必须留空", context.exception.errors[0].message)

    def test_video_with_custom_cover_requires_valid_ratio(self):
        for ratio in ("", "16:9", "original"):
            with self.subTest(ratio=ratio):
                content = self.build_workbook(
                    [[str(self.video), str(self.cover), ratio, "夏季女鞋穿搭", "", "", "", "", ""]]
                )
                with self.assertRaises(BatchValidationError) as context:
                    parse_tmall_video_batch_workbook(
                        content, account="shop1", dry_run=True, headed=True
                    )
                self.assertIn("封面比例", context.exception.errors[0].message)

    def test_tmall_article_batch_accepts_chinese_commas_in_tags(self):
        image = self.base_dir / "photos" / "001.jpg"
        image.write_bytes(b"image")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["图片文件夹路径", "封面比例", "标题", "发布文案", "标签"])
        worksheet.append([str(self.base_dir / "photos"), "3:4", "夏季女鞋图文", "轻盈舒适", "女鞋，夏季穿搭,通勤鞋"])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        rows = parse_tmall_article_batch_workbook(
            output.getvalue(),
            account="shop1",
            dry_run=True,
            headed=True,
        )

        self.assertEqual(rows[0].request.tags, ("女鞋", "夏季穿搭", "通勤鞋"))

    def test_tmall_article_rows_map_each_selected_cover_ratio(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["图片文件夹路径", "封面比例", "标题"])
        for ratio in ("原始", "3:4", "1:1"):
            worksheet.append([str(self.base_dir / "photos"), ratio, f"{ratio} 图文"])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        rows = parse_tmall_article_batch_workbook(
            output.getvalue(), account="shop1", dry_run=True, headed=True
        )

        self.assertEqual(
            [row.request.cover_ratio for row in rows], ["original", "3:4", "1:1"]
        )

    def test_tmall_article_requires_valid_cover_ratio(self):
        for ratio in ("", "16:9", "original"):
            with self.subTest(ratio=ratio):
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.append(["图片文件夹路径", "封面比例", "标题"])
                worksheet.append([str(self.base_dir / "photos"), ratio, "夏季女鞋图文"])
                output = BytesIO()
                workbook.save(output)
                workbook.close()
                with self.assertRaises(BatchValidationError) as context:
                    parse_tmall_article_batch_workbook(
                        output.getvalue(), account="shop1", dry_run=True, headed=True
                    )
                self.assertIn("封面比例", context.exception.errors[0].message)

    def test_explicit_creator_declaration_maps_to_request(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["视频路径", "自定义封面", "封面比例", "标题", "创作者声明"])
        worksheet.append([str(self.video), "", "", "夏季女鞋穿搭", "内容含营销广告"])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        rows = parse_tmall_video_batch_workbook(
            output.getvalue(),
            account="shop1",
            dry_run=True,
            headed=True,
        )

        self.assertEqual(rows[0].request.creator_declaration, "内容含营销信息")

    def test_music_name_maps_to_tmall_publish_request(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["视频路径", "自定义封面", "封面比例", "标题", "音乐名称", "创作者声明"])
        worksheet.append([str(self.video), "", "", "夏季女鞋穿搭", "默契", "内容无需标注"])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        rows = parse_tmall_video_batch_workbook(
            output.getvalue(),
            account="shop1",
            dry_run=True,
            headed=True,
        )

        self.assertEqual(rows[0].request.music_name, "默契")

    def test_packaged_video_template_includes_conditional_cover_ratio(self):
        from webapp.api.batch_templates import build_batch_template
        workbook = load_workbook(BytesIO(build_batch_template("tmall")))
        try:
            worksheet = workbook.active
            self.assertEqual(
                [worksheet.cell(1, column).value for column in range(1, 12)],
                [
                    "视频路径",
                    "自定义封面",
                    "封面比例",
                    "标题",
                    "文案",
                    "标签",
                    "商品ID",
                    "活动话题",
                    "音乐名称",
                    "定时发布",
                    "创作者声明",
                ],
            )
            self.assertIsNone(worksheet["B2"].value)
            self.assertIsNone(worksheet["C2"].value)
            self.assertEqual(worksheet["I2"].value, "默契")
            self.assertFalse(list(worksheet.merged_cells.ranges))
            validations = worksheet.data_validations.dataValidation
            self.assertEqual(len(validations), 2)
            validations_by_range = {str(item.sqref): item for item in validations}
            self.assertEqual(set(validations_by_range), {"C2:C201", "K2:K201"})
            ratio_formula = validations_by_range["C2:C201"].formula1
            self.assertIn('$B2<>""', ratio_formula)
            self.assertIn("BatchValidationOptions1", ratio_formula)
            self.assertIn("BatchValidationBlank1", ratio_formula)
            self.assertEqual(workbook["_validation_options"].sheet_state, "hidden")
            self.assertIn("BatchValidationOptions1", workbook.defined_names)
            self.assertIn("BatchValidationBlank1", workbook.defined_names)
        finally:
            workbook.close()

    def test_packaged_article_template_includes_cover_ratio_dropdown(self):
        from webapp.api.batch_templates import build_batch_template
        workbook = load_workbook(BytesIO(build_batch_template("tmall", "article")))
        try:
            worksheet = workbook.active
            self.assertEqual(worksheet["A1"].value, "图片文件夹路径")
            self.assertEqual(worksheet["B1"].value, "封面比例")
            self.assertEqual(worksheet["B2"].value, "3:4")
            validations = {
                str(item.sqref): item for item in worksheet.data_validations.dataValidation
            }
            self.assertEqual(validations["B2:B201"].formula1, '"原始,3:4,1:1"')
        finally:
            workbook.close()

    def test_blank_creator_declaration_is_rejected_when_column_exists(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["视频路径", "自定义封面", "封面比例", "标题", "创作者声明"])
        worksheet.append([str(self.video), "", "", "夏季女鞋穿搭", ""])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        with self.assertRaises(BatchValidationError) as context:
            parse_tmall_video_batch_workbook(
                output.getvalue(),
                account="shop1",
                dry_run=True,
                headed=True,
            )

        self.assertEqual(context.exception.errors[0].row, 2)
        self.assertIn("创作者声明", context.exception.errors[0].message)

    def test_template_intro_rows_are_skipped_before_the_header(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["天猫光合批量发布导入模板"])
        worksheet.append(["店铺账号在网页中选择"])
        worksheet.append([])
        worksheet.append(["视频路径", "自定义封面", "封面比例", "标题", "文案", "标签", "商品ID", "活动话题", "定时发布", "创作者声明"])
        worksheet.append([str(self.video), "", "", "夏季女鞋穿搭", "", "", "", "", "", "内容无需标注"])
        content = BytesIO()
        workbook.save(content)
        workbook.close()

        rows = parse_tmall_video_batch_workbook(
            content.getvalue(),
            account="shop1",
            dry_run=True,
            headed=True,
        )

        self.assertEqual(rows[0].row_number, 5)

    def test_invalid_rows_report_excel_row_without_creating_requests(self):
        content = self.build_workbook([[str(self.base_dir / "missing.mp4"), "", "", "", "", "", "", "", ""]])

        with self.assertRaises(BatchValidationError) as context:
            parse_tmall_video_batch_workbook(
                content,
                account="shop1",
                dry_run=False,
                headed=True,
            )

        self.assertEqual(context.exception.errors[0].row, 2)
        self.assertIn("视频文件不存在", context.exception.errors[0].message)

    def test_parent_directory_escape_is_reported_as_a_row_error(self):
        content = self.build_workbook([["demo.mp4", "", "", "夏季女鞋穿搭", "", "", "", "", ""]])

        with self.assertRaises(BatchValidationError) as context:
            parse_tmall_video_batch_workbook(
                content,
                account="shop1",
                dry_run=True,
                headed=True,
            )

        self.assertEqual(context.exception.errors[0].row, 2)
        self.assertIn("视频路径必须填写本机绝对路径", context.exception.errors[0].message)


class JdBatchWorkbookTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.base_dir = Path(self.temp_dir.name)
        self.video = self.base_dir / "photos" / "demo.mp4"
        self.video.parent.mkdir()
        self.video.write_bytes(b"video")

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def build_workbook(self, rows: list[list[object]]) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["视频路径", "标题", "商品ID", "定时发布", "自主原创", "创作者声明"])
        for row in rows:
            worksheet.append([*row, "内容无需标注"])
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    def test_valid_rows_map_to_jd_publish_requests(self):
        content = self.build_workbook(
            [[str(self.video), "京东视频标题示例", "12345", "", "是"]]
        )

        rows = parse_jd_video_batch_workbook(
            content,
            account="shop1",
            dry_run=True,
            headed=True,
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].row_number, 2)
        self.assertEqual(rows[0].request.platform, "jd")
        self.assertEqual(rows[0].request.video_path, self.video.resolve())
        self.assertEqual(rows[0].request.goods_id, "12345")
        self.assertTrue(rows[0].request.original)
        self.assertIsNone(rows[0].request.cover_image_path)

    def test_explicit_creator_declaration_maps_to_request(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["视频路径", "标题", "创作者声明"])
        worksheet.append([str(self.video), "京东视频标题示例", "含AI生成内容"])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        rows = parse_jd_video_batch_workbook(
            output.getvalue(),
            account="shop1",
            dry_run=True,
            headed=True,
        )

        self.assertEqual(rows[0].request.creator_declaration, "含AI生成内容")

    def test_blank_creator_declaration_is_rejected_when_column_exists(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["视频路径", "标题", "创作者声明"])
        worksheet.append([str(self.video), "京东视频标题示例", ""])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        with self.assertRaises(BatchValidationError) as context:
            parse_jd_video_batch_workbook(
                output.getvalue(),
                account="shop1",
                dry_run=True,
                headed=True,
            )

        self.assertEqual(context.exception.errors[0].row, 2)
        self.assertIn("创作者声明", context.exception.errors[0].message)

    def test_invalid_original_value_reports_its_excel_column(self):
        content = self.build_workbook(
            [[str(self.video), "京东视频标题示例", "", "", "不确定"]]
        )

        with self.assertRaises(BatchValidationError) as context:
            parse_jd_video_batch_workbook(
                content,
                account="shop1",
                dry_run=False,
                headed=True,
            )

        self.assertEqual(context.exception.errors[0].row, 2)
        self.assertEqual(context.exception.errors[0].field, "自主原创")


class SocialBatchWorkbookTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.base_dir = Path(self.temp_dir.name)
        self.video = self.base_dir / "demo.mp4"
        self.video.write_bytes(b"video")
        self.image_dir = self.base_dir / "images"
        self.image_dir.mkdir()
        (self.image_dir / "001.jpg").write_bytes(b"image")

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def workbook(self, header: list[str], row: list[object]) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(header)
        worksheet.append(row)
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    def test_social_video_workbooks_map_to_publish_requests(self):
        content = self.workbook(
            ["视频路径", "标题", "视频描述", "标签"],
            [str(self.video), "抖音标题", "描述", "热点，穿搭"],
        )
        rows = parse_douyin_video_batch_workbook(
            content, account="shop1", dry_run=True, headed=False
        )
        self.assertEqual(rows[0].request.platform, "douyin")
        self.assertEqual(rows[0].request.tags, ("热点", "穿搭"))
        self.assertTrue(rows[0].request.dry_run)
        self.assertIsNone(rows[0].request.cover_image_path)

        content = self.workbook(
            ["视频路径", "标题", "笔记正文", "标签"],
            [str(self.video), "小红书标题", "正文", "种草，穿搭"],
        )
        rows = parse_xiaohongshu_video_batch_workbook(
            content, account="shop1", dry_run=True, headed=True
        )
        self.assertEqual(rows[0].request.platform, "xiaohongshu")
        self.assertEqual(rows[0].request.tags, ("种草", "穿搭"))
        self.assertIsNone(rows[0].request.cover_image_path)

    def test_social_article_workbooks_map_to_publish_requests(self):
        xhs_rows = parse_xiaohongshu_article_batch_workbook(
            self.workbook(
                ["图片文件夹路径", "标题", "笔记正文", "标签"],
                [str(self.image_dir), "小红书图文", "正文", "种草，穿搭"],
            ),
            account="shop1",
            dry_run=True,
            headed=True,
        )
        self.assertEqual(xhs_rows[0].request.platform, "xiaohongshu")
        self.assertEqual(
            xhs_rows[0].request.image_paths,
            ((self.image_dir / "001.jpg").resolve(),),
        )
        self.assertEqual(xhs_rows[0].request.tags, ("种草", "穿搭"))

        douyin_rows = parse_douyin_article_batch_workbook(
            self.workbook(
                ["图片文件夹路径", "标题", "图文描述", "标签"],
                [str(self.image_dir), "抖音图文", "描述", "热点，穿搭"],
            ),
            account="shop1",
            dry_run=True,
            headed=False,
        )
        self.assertEqual(douyin_rows[0].request.platform, "douyin")
        self.assertEqual(douyin_rows[0].request.description, "描述")
        self.assertEqual(douyin_rows[0].request.tags, ("热点", "穿搭"))

    def test_batch_template_dispatch_includes_social_platforms(self):
        cases = {
            ("xiaohongshu", "video"): ["视频路径", "自定义封面", "标题", "笔记正文", "标签", "定时发布"],
            ("xiaohongshu", "article"): ["图片文件夹路径", "标题", "笔记正文", "标签", "定时发布"],
            ("douyin", "video"): ["视频路径", "横版封面", "标题", "视频描述", "标签", "定时发布"],
            ("douyin", "article"): ["图片文件夹路径", "标题", "图文描述", "标签", "定时发布"],
        }
        for (platform, content_type), expected_headers in cases.items():
            workbook = load_workbook(BytesIO(build_batch_template(platform, content_type)))
            try:
                worksheet = workbook.active
                headers = [
                    worksheet.cell(1, column).value
                    for column in range(1, len(expected_headers) + 1)
                ]
                self.assertEqual(headers, expected_headers)
            finally:
                workbook.close()


class TmallBatchApiTests(unittest.TestCase):
    def test_valid_workbook_creates_one_job_per_excel_row(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            media_dir = AppDataPaths.create(
                root / ".runtime-test-users"
            ).for_user(TEST_USER_ID).media
            video = media_dir / "demo.mp4"
            video.write_bytes(b"video")
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["视频路径", "自定义封面", "封面比例", "标题", "文案", "标签", "商品ID", "活动话题", "定时发布", "创作者声明"])
            worksheet.append([str(video), "", "", "夏季女鞋穿搭", "轻盈舒适", "女鞋,夏季穿搭", "12345", "", "", "内容无需标注"])
            worksheet.append([str(video), "", "", "夏季通勤穿搭", "舒适百搭", "通勤", "", "", "", "内容无需标注"])
            content = BytesIO()
            workbook.save(content)
            workbook.close()

            store = JobStore(root / "runtime")
            manager = TaskManager(store, runner=lambda job: {"message": "complete"}, max_workers=1)
            app = create_app(
                WebSettings(data_dir=root / "app-data", frontend_dist_dir=root / "missing"),
                manager,
            )
            endpoint = next(route.endpoint for route in app.routes if route.path == "/api/jobs/batch/tmall")
            upload = UploadFile(filename="tmall.xlsx", file=BytesIO(content.getvalue()))
            try:
                response = asyncio.run(
                    endpoint(
                        account="shop1",
                        workbook=upload,
                        dry_run=True,
                        headed=True,
                        workspace=app.state.test_workspace,
                    )
                )
                body = json.loads(response.body)

                self.assertEqual(response.status_code, 202)
                self.assertEqual(body["created_count"], 2)
                self.assertEqual([job["source_row"] for job in body["jobs"]], [2, 3])
                self.assertTrue(all(job["batch_id"] == body["batch_id"] for job in body["jobs"]))
                for _ in range(50):
                    statuses = [store.get_job(job["id"])["status"] for job in body["jobs"]]
                    if all(status == "succeeded" for status in statuses):
                        break
                    time.sleep(0.02)
                else:
                    self.fail("batch jobs did not complete")
            finally:
                manager.shutdown()

    def test_invalid_batch_does_not_create_partial_jobs(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            media_dir = AppDataPaths.create(
                root / ".runtime-test-users"
            ).for_user(TEST_USER_ID).media
            video = media_dir / "demo.mp4"
            video.write_bytes(b"video")
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["视频路径", "自定义封面", "封面比例", "标题", "创作者声明"])
            worksheet.append([str(video), "", "", "有效标题", "内容无需标注"])
            worksheet.append([str(root / "missing.mp4"), "", "", "无效视频", "内容无需标注"])
            content = BytesIO()
            workbook.save(content)
            workbook.close()

            store = JobStore(root / "runtime")
            manager = TaskManager(store, runner=lambda job: {"message": "complete"}, max_workers=1)
            app = create_app(
                WebSettings(data_dir=root / "app-data", frontend_dist_dir=root / "missing"),
                manager,
            )
            endpoint = next(route.endpoint for route in app.routes if route.path == "/api/jobs/batch/tmall")
            upload = UploadFile(filename="tmall.xlsx", file=BytesIO(content.getvalue()))
            try:
                response = asyncio.run(
                    endpoint(
                        account="shop1",
                        workbook=upload,
                        dry_run=True,
                        headed=True,
                        workspace=app.state.test_workspace,
                    )
                )
                body = json.loads(response.body)

                self.assertEqual(response.status_code, 422)
                self.assertEqual(body["detail"], "Excel 内容校验失败，未创建任何发布任务")
                self.assertEqual(len(manager.store.list_jobs(limit=None)), 0)
            finally:
                manager.shutdown()


class JdBatchApiTests(unittest.TestCase):
    def test_valid_article_workbook_creates_jd_article_job(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            image_dir = root / "article-images"
            image_dir.mkdir()
            (image_dir / "01.jpg").write_bytes(b"image")
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append([
                "图片文件夹路径", "标题", "正文内容", "商品ID", "参与话题",
                "定时发布", "自主原创", "创作者声明",
            ])
            worksheet.append([
                str(image_dir), "京东图文标题", "京东图文正文", "12345", "数码先锋",
                "", "否", "内容无需标注",
            ])
            content = BytesIO()
            workbook.save(content)
            workbook.close()

            store = JobStore(root / "runtime")
            manager = TaskManager(store, runner=lambda job: {"message": "complete"}, max_workers=1)
            app = create_app(
                WebSettings(data_dir=root / "app-data", frontend_dist_dir=root / "missing"),
                manager,
            )
            endpoint = next(route.endpoint for route in app.routes if route.path == "/api/jobs/batch/jd")
            upload = UploadFile(filename="jd-article.xlsx", file=BytesIO(content.getvalue()))
            try:
                response = asyncio.run(
                    endpoint(
                        account="shop1",
                        workbook=upload,
                        content_type="article",
                        dry_run=True,
                        headed=True,
                        workspace=app.state.test_workspace,
                    )
                )
                body = json.loads(response.body)

                self.assertEqual(response.status_code, 202)
                self.assertEqual(body["created_count"], 1)
                created = store.get_job(body["jobs"][0]["id"])
                self.assertEqual(created["payload"]["content_type"], "article")
                self.assertEqual(created["payload"]["description"], "京东图文正文")
                self.assertEqual(created["payload"]["activity_topic"], "数码先锋")
            finally:
                manager.shutdown()

    def test_valid_workbook_creates_one_job_per_excel_row(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            media_dir = AppDataPaths.create(
                root / ".runtime-test-users"
            ).for_user(TEST_USER_ID).media
            video = media_dir / "demo.mp4"
            video.write_bytes(b"video")
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["视频路径", "标题", "商品ID", "定时发布", "自主原创", "创作者声明"])
            worksheet.append([str(video), "京东视频标题示例", "12345", "", "是", "内容无需标注"])
            worksheet.append([str(video), "京东夏日好物推荐", "", "", "否", "内容无需标注"])
            content = BytesIO()
            workbook.save(content)
            workbook.close()

            store = JobStore(root / "runtime")
            manager = TaskManager(store, runner=lambda job: {"message": "complete"}, max_workers=1)
            app = create_app(
                WebSettings(data_dir=root / "app-data", frontend_dist_dir=root / "missing"),
                manager,
            )
            endpoint = next(route.endpoint for route in app.routes if route.path == "/api/jobs/batch/jd")
            upload = UploadFile(filename="jd.xlsx", file=BytesIO(content.getvalue()))
            try:
                response = asyncio.run(
                    endpoint(
                        account="shop1",
                        workbook=upload,
                        dry_run=True,
                        headed=True,
                        workspace=app.state.test_workspace,
                    )
                )
                body = json.loads(response.body)

                self.assertEqual(response.status_code, 202)
                self.assertEqual(body["created_count"], 2)
                self.assertEqual([job["platform"] for job in body["jobs"]], ["jd", "jd"])
                self.assertEqual([job["source_row"] for job in body["jobs"]], [2, 3])
                self.assertTrue(all(job["batch_id"] == body["batch_id"] for job in body["jobs"]))
                for _ in range(50):
                    statuses = [store.get_job(job["id"])["status"] for job in body["jobs"]]
                    if all(status == "succeeded" for status in statuses):
                        break
                    time.sleep(0.02)
                else:
                    self.fail("batch jobs did not complete")
            finally:
                manager.shutdown()

    def test_invalid_batch_does_not_create_partial_jobs(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            media_dir = AppDataPaths.create(
                root / ".runtime-test-users"
            ).for_user(TEST_USER_ID).media
            video = media_dir / "demo.mp4"
            video.write_bytes(b"video")
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["视频路径", "标题", "自主原创", "创作者声明"])
            worksheet.append([str(video), "京东视频标题示例", "否", "内容无需标注"])
            worksheet.append([str(video), "京东夏日好物推荐", "未知", "内容无需标注"])
            content = BytesIO()
            workbook.save(content)
            workbook.close()

            store = JobStore(root / "runtime")
            manager = TaskManager(store, runner=lambda job: {"message": "complete"}, max_workers=1)
            app = create_app(
                WebSettings(data_dir=root / "app-data", frontend_dist_dir=root / "missing"),
                manager,
            )
            endpoint = next(route.endpoint for route in app.routes if route.path == "/api/jobs/batch/jd")
            upload = UploadFile(filename="jd.xlsx", file=BytesIO(content.getvalue()))
            try:
                response = asyncio.run(
                    endpoint(
                        account="shop1",
                        workbook=upload,
                        dry_run=True,
                        headed=True,
                        workspace=app.state.test_workspace,
                    )
                )
                body = json.loads(response.body)

                self.assertEqual(response.status_code, 422)
                self.assertEqual(body["errors"][0]["row"], 3)
                self.assertEqual(body["errors"][0]["field"], "自主原创")
                self.assertEqual(store.list_jobs(), [])
            finally:
                manager.shutdown()

class JobStoreTests(unittest.TestCase):
    def test_recovery_returns_queued_jobs_in_creation_order(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            jobs = [
                store.create_job(
                    kind="check", platform="tmall", account="shop1", payload={"index": index}
                )
                for index in range(3)
            ]
            id_order = sorted(jobs, key=lambda job: job["id"])
            expected = list(reversed(id_order))
            for index, job in enumerate(expected):
                store.update_job(job["id"], created_at=f"2026-01-01T00:00:0{index}+00:00")

            recovered = store.recover_interrupted_jobs()

            self.assertEqual(recovered, [job["id"] for job in expected])

    def test_batch_is_persisted_with_one_atomic_state_write(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = JobStore(root)
            definitions = [
                {
                    "kind": "publish",
                    "platform": "jd",
                    "account": "shop1",
                    "payload": {"title": f"item-{index}"},
                }
                for index in range(3)
            ]

            with patch.object(store, "_write", wraps=store._write) as write:
                jobs = store.create_jobs(definitions)

            self.assertEqual(len(jobs), 3)
            self.assertEqual(write.call_count, 1)
            self.assertEqual(len(store.list_jobs()), 3)
            self.assertEqual(root.stat().st_mode & 0o777, 0o700)
            self.assertEqual(store.path.stat().st_mode & 0o777, 0o600)

    def test_failed_batch_state_write_does_not_leave_partial_jobs(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            definitions = [
                {
                    "kind": "publish",
                    "platform": "tmall",
                    "account": "shop1",
                    "payload": {"title": f"item-{index}"},
                }
                for index in range(2)
            ]

            with patch.object(store, "_write", side_effect=OSError("disk full")):
                with self.assertRaisesRegex(OSError, "disk full"):
                    store.create_jobs(definitions)

            self.assertEqual(store.list_jobs(), [])

    def test_terminal_history_is_pruned_but_uncertain_jobs_are_retained(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            old = (datetime.now(timezone.utc) - timedelta(days=120)).isoformat()
            ordinary = store.create_job(
                kind="check", platform="tmall", account="shop1", payload={}
            )
            uncertain = store.create_job(
                kind="publish", platform="tmall", account="shop1", payload={}
            )
            store.update_job(
                ordinary["id"], status="succeeded", finished_at=old
            )
            store.update_job(
                uncertain["id"], status="uncertain", finished_at=old
            )

            removed = store.prune_terminal_jobs(max_count=10, older_than_days=90)

            self.assertEqual([job["id"] for job in removed], [ordinary["id"]])
            self.assertIsNotNone(store.get_job(uncertain["id"]))

    def test_job_list_supports_a_full_batch_and_reports_global_counts(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            for index in range(205):
                job = store.create_job(
                    kind="check", platform="tmall", account=f"shop{index}", payload={}
                )
                if index < 3:
                    store.update_job(job["id"], status="succeeded")

            self.assertEqual(len(store.list_jobs(limit=500)), 205)
            summary = store.job_summary()
            self.assertEqual(summary["total"], 205)
            self.assertEqual(summary["statuses"]["succeeded"], 3)

    def test_delete_job_only_allows_terminal_tasks(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            job = store.create_job(
                kind="check", platform="tmall", account="shop1", payload={"headed": False}
            )

            with self.assertRaisesRegex(ValueError, "已完成或失败"):
                store.delete_job(job["id"])

            store.update_job(job["id"], status="succeeded")
            deleted = store.delete_job(job["id"])

            self.assertEqual(deleted["id"], job["id"])
            self.assertIsNone(store.get_job(job["id"]))

    def test_delete_account_removes_only_the_saved_account(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            store.remember_account("tmall", "shop1")
            store.remember_account("jd", "shop1")

            deleted = store.delete_account("tmall", "shop1")

            self.assertEqual(deleted["platform"], "tmall")
            remaining_accounts = store.list_accounts()
            self.assertEqual(len(remaining_accounts), 1)
            self.assertEqual(remaining_accounts[0]["platform"], "jd")
            self.assertEqual(remaining_accounts[0]["account"], "shop1")

    def test_delete_account_rejects_an_account_with_active_tasks(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = JobStore(Path(temp_dir))
            store.remember_account("tmall", "shop1")
            store.create_job(
                kind="login", platform="tmall", account="shop1", payload={"headed": True}
            )

            with self.assertRaisesRegex(ValueError, "排队中或执行中的任务"):
                store.delete_account("tmall", "shop1")


class ApiEndpointTests(unittest.TestCase):
    def test_delete_account_removes_cookie_and_dropdown_entry(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = JobStore(root / "runtime")
            store.remember_account("tmall", "shop1")
            manager = TaskManager(store, runner=lambda job: {"message": "complete"}, max_workers=1)
            app = create_app(
                WebSettings(data_dir=root / "app-data", frontend_dist_dir=root / "missing"),
                manager,
            )
            endpoint = next(
                route.endpoint for route in app.routes if route.path == "/api/accounts/{platform}/{account}" and "DELETE" in route.methods
            )
            try:
                with patch.object(manager, "close_account_session") as close_session, patch(
                    "webapp.api.main.delete_account_cookie", return_value=True
                ) as delete_cookie:
                    response = endpoint(
                        "tmall", "shop1", workspace=app.state.test_workspace
                    )
                body = response

                self.assertTrue(body["cookie_deleted"])
                close_session.assert_called_once_with("tmall", "shop1")
                delete_cookie.assert_called_once_with(manager.paths, "tmall", "shop1")
                self.assertEqual(store.list_accounts(), [])
            finally:
                manager.shutdown()
    def test_log_delete_failure_keeps_the_task_record(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = JobStore(root / "runtime")
            job = store.create_job(
                kind="check", platform="tmall", account="shop1", payload={}
            )
            store.update_job(job["id"], status="succeeded")
            manager = TaskManager(store, runner=lambda _job: {"message": "complete"})
            app = create_app(
                WebSettings(data_dir=root / "app-data", frontend_dist_dir=root / "missing"),
                manager,
            )
            endpoint = next(
                route.endpoint
                for route in app.routes
                if route.path == "/api/jobs/{job_id}" and "DELETE" in route.methods
            )
            try:
                with patch.object(
                    manager,
                    "delete_job_artifacts",
                    side_effect=OSError("permission denied"),
                ):
                    with self.assertRaises(HTTPException) as context:
                        endpoint(job["id"], workspace=app.state.test_workspace)

                self.assertEqual(context.exception.status_code, 500)
                self.assertIsNotNone(store.get_job(job["id"]))
            finally:
                manager.shutdown()

    def test_batch_delete_works_with_local_agent_manager(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            paths = AppDataPaths.create(root / "data").for_user(TEST_USER_ID)
            store = JobStore(paths.runtime)
            jobs = [
                store.create_job(
                    kind="check",
                    platform="tmall",
                    account=f"shop{index}",
                    payload={},
                )
                for index in range(2)
            ]
            for job in jobs:
                store.update_job(job["id"], status="succeeded")
            manager = AgentTaskManager(store, user_id=TEST_USER_ID, paths=paths)
            app = create_app(
                WebSettings(data_dir=root / "app-data", frontend_dist_dir=root / "missing"),
                manager,
            )
            endpoint = next(
                route.endpoint
                for route in app.routes
                if route.path == "/api/jobs/batch-delete"
            )
            try:
                for job in jobs:
                    manager.job_log_path(job["id"]).write_text("local log", encoding="utf-8")

                body = endpoint(
                    {"job_ids": [job["id"] for job in jobs]},
                    workspace=app.state.test_workspace,
                )

                self.assertEqual(body["deleted"], [job["id"] for job in jobs])
                self.assertEqual(body["skipped"], [])
                self.assertTrue(all(store.get_job(job["id"]) is None for job in jobs))
                self.assertTrue(all(not manager.job_log_path(job["id"]).exists() for job in jobs))
            finally:
                manager.shutdown()

    def test_cancel_queued_task_preserves_its_account(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = JobStore(root / "runtime")
            release_worker = threading.Event()
            worker_started = threading.Event()

            def block_worker(_job):
                worker_started.set()
                release_worker.wait(timeout=2)
                return {"message": "complete"}

            manager = TaskManager(store, runner=block_worker, max_workers=1)
            first_job = manager.submit_account_task(
                kind="check", platform="tmall", account="shop1", headed=False
            )
            self.assertTrue(worker_started.wait(timeout=1))
            cancelled_job = manager.submit_account_task(
                kind="login", platform="tmall", account="shop2", headed=True
            )
            app = create_app(
                WebSettings(data_dir=root / "app-data", frontend_dist_dir=root / "missing"),
                manager,
            )
            endpoint = next(
                route.endpoint
                for route in app.routes
                if route.path == "/api/jobs/{job_id}/cancel"
            )
            try:
                with patch("webapp.api.main.delete_account_cookie", return_value=True) as delete_cookie:
                    result = endpoint(
                        cancelled_job["id"],
                        workspace=app.state.test_workspace,
                    )

                self.assertEqual(result["job"]["status"], "cancelled")
                delete_cookie.assert_not_called()
                self.assertEqual(store.get_job(first_job["id"])["status"], "running")
                self.assertEqual(
                    [(item["platform"], item["account"]) for item in store.list_accounts()],
                    [("tmall", "shop1"), ("tmall", "shop2")],
                )
            finally:
                release_worker.set()
                for _ in range(50):
                    completed = store.get_job(first_job["id"])
                    if completed and completed["status"] == "succeeded":
                        break
                    time.sleep(0.02)
                manager.shutdown()

    def test_batch_cancel_interrupts_active_jobs_and_skips_other_ids(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = JobStore(root / "runtime")
            manager = AgentTaskManager(
                store,
                user_id=TEST_USER_ID,
                paths=AppDataPaths.create(root / "data").for_user(TEST_USER_ID),
            )
            queued_jobs = [
                manager.submit_account_task(
                    kind="check", platform="tmall", account=f"shop{index}", headed=False
                )
                for index in range(2)
            ]
            completed_job = store.create_job(
                kind="check", platform="tmall", account="finished", payload={}
            )
            store.update_job(completed_job["id"], status="succeeded")
            app = create_app(
                WebSettings(data_dir=root / "app-data", frontend_dist_dir=root / "missing"),
                manager,
            )
            endpoint = next(
                route.endpoint
                for route in app.routes
                if route.path == "/api/jobs/batch-cancel"
            )
            try:
                result = endpoint(
                    {
                        "job_ids": [
                            queued_jobs[0]["id"],
                            queued_jobs[1]["id"],
                            completed_job["id"],
                            "missing-job",
                        ]
                    },
                    workspace=app.state.test_workspace,
                )

                self.assertEqual(
                    result["cancelled"], [job["id"] for job in queued_jobs]
                )
                self.assertEqual(len(result["skipped"]), 2)
                self.assertTrue(
                    all(store.get_job(job["id"])["status"] == "cancelled" for job in queued_jobs)
                )
                self.assertEqual(store.get_job(completed_job["id"])["status"], "succeeded")
                self.assertEqual(len(store.list_accounts()), 3)
            finally:
                manager.shutdown()


class ReadinessTests(unittest.TestCase):
    @staticmethod
    def readiness_endpoint(app):
        return next(route.endpoint for route in app.routes if route.path == "/api/readiness")

    def test_partial_frontend_is_reported_without_mounting_missing_assets(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            frontend = root / "frontend"
            frontend.mkdir()
            (frontend / "index.html").write_text("partial", encoding="utf-8")
            store = JobStore(root / "runtime")
            manager = TaskManager(store, runner=lambda _job: {"message": "complete"})

            app = create_app(WebSettings(root / "app-data", frontend), manager)
            manager.start()
            try:
                response = self.readiness_endpoint(app)()
                body = json.loads(response.body)

                self.assertEqual(response.status_code, 503)
                self.assertFalse(body["checks"]["frontend_built"])
                self.assertTrue(body["checks"]["workspace_registry"])
                self.assertEqual(
                    body["capacity"],
                    {
                        "active_jobs_per_agent": 1,
                        "browser_capacity_location": "user_device",
                    },
                )
                root_endpoint = next(
                    route.endpoint for route in app.routes if getattr(route, "path", None) == "/"
                )
                self.assertIn("请在 webapp/frontend", root_endpoint()["message"])
            finally:
                manager.shutdown()

    def test_ready_installation_degrades_when_maintenance_fails(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            frontend = root / "frontend"
            (frontend / "assets").mkdir(parents=True)
            (frontend / "index.html").write_text("ready", encoding="utf-8")
            store = JobStore(root / "runtime")
            manager = TaskManager(store, runner=lambda _job: {"message": "complete"})
            app = create_app(WebSettings(root / "app-data", frontend), manager)
            manager.start()
            try:
                ready = self.readiness_endpoint(app)()
                manager._maintenance_errors.append("cleanup failed")
                degraded = self.readiness_endpoint(app)()

                self.assertEqual(ready.status_code, 200)
                self.assertEqual(degraded.status_code, 503)
                body = json.loads(degraded.body)
                self.assertFalse(body["checks"]["maintenance_clean"])
                self.assertEqual(
                    body["maintenance_errors"],
                    [f"{TEST_USER_ID}: cleanup failed"],
                )
            finally:
                manager.shutdown()


class PlatformAdapterTests(unittest.TestCase):
    def test_cookie_directory_permissions_are_restricted(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            paths = AppDataPaths.create(Path(temp_dir) / "data").for_user(
                TEST_USER_ID
            )
            account_file = resolve_account_file(paths, "tmall", "shop1")

            self.assertEqual(account_file.parent.stat().st_mode & 0o777, 0o700)

    def test_cookie_permissions_are_restricted_to_the_current_user(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            cookie_file = Path(temp_dir) / "tmall_shop1.json"
            cookie_file.write_text("[]", encoding="utf-8")

            secure_account_file(cookie_file)

            self.assertEqual(cookie_file.stat().st_mode & 0o777, 0o600)

    def test_delete_account_cookie_removes_only_its_file(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            paths = AppDataPaths.create(Path(temp_dir) / "data").for_user(
                TEST_USER_ID
            )
            cookie_file = Path(temp_dir) / "tmall_shop1.json"
            cookie_file.write_text("[]", encoding="utf-8")

            with patch("webapp.api.platforms.resolve_account_file", return_value=cookie_file):
                self.assertTrue(delete_account_cookie(paths, "tmall", "shop1"))
                self.assertFalse(cookie_file.exists())
                self.assertFalse(delete_account_cookie(paths, "tmall", "shop1"))

    def test_tmall_publish_adapter_calls_pooled_uploader(self):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        paths = AppDataPaths.create(Path(temp_dir.name) / "data").for_user(
            TEST_USER_ID
        )
        request = TmallVideoUploadRequest(
            account_name="shop1",
            video_file=Path("/tmp/demo.mp4"),
            cover_ratio="original",
            title="夏季女鞋测评",
            description="轻便好穿",
            tags=["女鞋"],
            goods_id="12345,67890",
            music_name="默契",
        )
        account_file = Path("/tmp/tmall_shop1.json")
        leased_session = object()

        class Lease:
            async def __aenter__(self):
                return leased_session

            async def __aexit__(self, _exc_type, _exc, _traceback):
                return False

        class Pool:
            def lease(self, _path, *, headless):
                self.headless = headless
                return Lease()

        pool = Pool()

        with patch("webapp.api.platforms.resolve_account_file", return_value=account_file), patch(
            "webapp.api.platforms.tmall_setup", new=AsyncMock(return_value=True)
        ), patch("webapp.api.platforms.TmallVideo") as uploader_type:
            uploader_type.return_value.upload_in_session = AsyncMock()
            result = asyncio.run(
                upload_tmall_video(request, paths=paths, session_pool=pool)
            )

        self.assertEqual(result, {})
        uploader_type.return_value.upload_in_session.assert_awaited_once_with(leased_session)
        self.assertEqual(uploader_type.call_args.kwargs["account_file"], str(account_file))
        self.assertEqual(uploader_type.call_args.kwargs["goods_id"], "12345,67890")
        self.assertEqual(uploader_type.call_args.kwargs["music_name"], "默契")
        self.assertNotIn("screenshot_dir", uploader_type.call_args.kwargs)

    def test_tmall_publish_adapter_passes_optional_cover_to_uploader(self):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        paths = AppDataPaths.create(Path(temp_dir.name) / "data").for_user(
            TEST_USER_ID
        )
        request = TmallVideoUploadRequest(
            account_name="shop1",
            video_file=Path("/tmp/demo.mp4"),
            cover_ratio="3:4",
            cover_image_file=Path("/tmp/cover.png"),
            title="夏季女鞋测评",
            description="轻便好穿",
            tags=[],
        )
        account_file = Path("/tmp/tmall_shop1.json")
        leased_session = object()

        class Lease:
            async def __aenter__(self):
                return leased_session

            async def __aexit__(self, _exc_type, _exc, _traceback):
                return False

        class Pool:
            def lease(self, _path, *, headless):
                return Lease()

        with patch(
            "webapp.api.platforms.resolve_account_file", return_value=account_file
        ), patch(
            "webapp.api.platforms.tmall_setup", new=AsyncMock(return_value=True)
        ), patch("webapp.api.platforms.TmallVideo") as uploader_type:
            uploader_type.return_value.upload_in_session = AsyncMock()
            asyncio.run(upload_tmall_video(request, paths=paths, session_pool=Pool()))

        self.assertEqual(
            uploader_type.call_args.kwargs["cover_image_path"], "/tmp/cover.png"
        )

    def test_jd_publish_adapter_calls_pooled_uploader(self):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        paths = AppDataPaths.create(Path(temp_dir.name) / "data").for_user(
            TEST_USER_ID
        )
        request = JdVideoUploadRequest(
            account_name="shop1",
            video_file=Path("/tmp/demo.mp4"),
            title="京东视频标题示例",
            goods_id="12345",
            original=True,
        )
        account_file = Path("/tmp/jd_shop1.json")
        leased_session = object()

        class Lease:
            async def __aenter__(self):
                return leased_session

            async def __aexit__(self, _exc_type, _exc, _traceback):
                return False

        class Pool:
            def lease(self, _path, *, headless):
                self.headless = headless
                return Lease()

        pool = Pool()

        with patch("webapp.api.platforms.resolve_account_file", return_value=account_file), patch(
            "webapp.api.platforms.jd_setup", new=AsyncMock(return_value=True)
        ), patch("webapp.api.platforms.JDVideo") as uploader_type:
            uploader_type.return_value.upload_in_session = AsyncMock()
            result = asyncio.run(
                upload_jd_video(request, paths=paths, session_pool=pool)
            )

        self.assertEqual(result, {})
        uploader_type.return_value.upload_in_session.assert_awaited_once_with(leased_session)
        self.assertEqual(uploader_type.call_args.kwargs["account_file"], str(account_file))
        self.assertNotIn("screenshot_dir", uploader_type.call_args.kwargs)


class LocalSecurityMiddlewareTests(unittest.TestCase):
    @staticmethod
    async def request_status(app, method: str, path: str, headers: dict[str, str]) -> int:
        messages: list[dict] = []
        request_sent = False

        async def receive():
            nonlocal request_sent
            if request_sent:
                return {"type": "http.disconnect"}
            request_sent = True
            return {"type": "http.request", "body": b"", "more_body": False}

        async def send(message):
            messages.append(message)

        scope = {
            "type": "http",
            "asgi": {"version": "3.0", "spec_version": "2.3"},
            "http_version": "1.1",
            "method": method,
            "scheme": "http",
            "path": path,
            "raw_path": path.encode("ascii"),
            "query_string": b"",
            "root_path": "",
            "headers": [
                (name.lower().encode("ascii"), value.encode("ascii"))
                for name, value in headers.items()
            ],
            "client": ("127.0.0.1", 12345),
            "server": ("testserver", 80),
        }
        await app(scope, receive, send)
        return next(message["status"] for message in messages if message["type"] == "http.response.start")

    def test_cross_site_mutation_and_untrusted_host_are_rejected(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = JobStore(root / "runtime")
            manager = TaskManager(store, runner=lambda _job: {"message": "complete"})
            app = create_app(
                WebSettings(data_dir=root / "app-data", frontend_dist_dir=root / "missing"),
                manager,
            )
            try:
                cross_site_status = asyncio.run(
                    self.request_status(
                        app,
                        "POST",
                        "/api/accounts/tmall/shop1/login",
                        {"host": "testserver", "origin": "https://attacker.example"},
                    )
                )
                untrusted_host_status = asyncio.run(
                    self.request_status(
                        app, "GET", "/api/health", {"host": "attacker.example"}
                    )
                )
                same_origin_status = asyncio.run(
                    self.request_status(
                        app,
                        "POST",
                        "/api/accounts/tmall/shop1/login",
                        {"host": "127.0.0.1:8788", "origin": "http://127.0.0.1:8788"},
                    )
                )

                self.assertEqual(cross_site_status, 403)
                self.assertEqual(untrusted_host_status, 400)
                self.assertIn(same_origin_status, {401, 403})
                self.assertEqual(len(store.list_jobs()), 0)
            finally:
                manager.shutdown()

    def test_new_task_without_its_own_log_does_not_show_another_task_platform_log(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = JobStore(root / "runtime")
            manager = TaskManager(
                store,
                runner=lambda _job: {"message": "complete"},
                job_log_dir=root / "job-logs",
            )
            job = store.create_job(
                kind="check", platform="tmall", account="shop1", payload={}
            )
            app = create_app(
                WebSettings(data_dir=root / "app-data", frontend_dist_dir=root / "missing"),
                manager,
            )
            endpoint = next(
                route.endpoint for route in app.routes if route.path == "/api/jobs/{job_id}"
            )
            try:
                with patch("webapp.api.main._tail_platform_log") as platform_log:
                    result = endpoint(job["id"], workspace=app.state.test_workspace)

                self.assertEqual(result["logs"], [])
                platform_log.assert_not_called()
            finally:
                manager.shutdown()

if __name__ == "__main__":
    unittest.main()
